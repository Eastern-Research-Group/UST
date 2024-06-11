import sys
import os
sys.path = [os.path.join(os.path.dirname(__file__), "..", "..")] + sys.path
from ust.python.util.logger_factory import logger
from ust.python.util import utils, config
from datetime import date
import openpyxl as op
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side


ref_heading_color = 'C9C9C9'

fill_gen = PatternFill(fill_type='solid',
					   start_color=ref_heading_color,
					   end_color=ref_heading_color)

thin_border = Border(left=Side(style='thin'), 
					 right=Side(style='thin'), 
					 top=Side(style='thin'), 
					 bottom=Side(style='thin'))

left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)



def get_headers(cur, view_name):
	sql = """select column_name from information_schema.columns
	         where table_schema = 'public' and table_name = %s
			 order by ordinal_position"""
	cur.execute(sql, (view_name, ))
	return [x[0] for x in cur.fetchall()]


def make_data_page(organization_id, ust_or_lust, wb):
	ws_name = organization_id + ' ' + ust_or_lust.upper()
	ws = wb.active
	ws.title = ws_name
	
	conn = utils.connect_db()
	cur = conn.cursor()

	view_name = 'v_' + ust_or_lust.lower() 
	headers = get_headers(cur, view_name)

	for colno, header in enumerate(headers, start=1):
		cell = ws.cell(row=1, column=colno)
		cell.value = header
		cell.font = Font(bold=True)

	sql = "select * from public." + view_name + ' where organization_id = %s'
	if ust_or_lust.lower() == 'ust':
		sql = sql + ' order by "FacilityID", "TankID", "CompartmentID"'
	else:
		sql = sql + ' order by "FacilityID", "SiteName", "LUSTID"'
	cur.execute(sql, (organization_id, ))
	data = cur.fetchall()

	for rowno, row in enumerate(data, start=2):
		for colno, cell_value in enumerate(row, start=1):
			ws.cell(row=rowno, column=colno).value = cell_value

	cur.close()
	conn.close()

	ws.delete_cols(1)
	ws.freeze_panes = ws['A2']

	logger.info('Created data page')


def get_reference_data(ust_or_lust, ws):
	conn = utils.connect_db()
	cur = conn.cursor()

	view_name = 'v_' + ust_or_lust.lower() + '_elements'
	headers = get_headers(cur, view_name)

	for colno, header in enumerate(headers, start=1):
		ws.cell(row=1, column=colno).value = header

	sql = "select * from public.v_" + ust_or_lust.lower() + '_elements'	
	cur.execute(sql)
	data = cur.fetchall()

	for rowno, row in enumerate(data, start=2):
		for colno, cell_value in enumerate(row, start=1):
			if colno == 7 and cell_value:
				cell_value = cell_value.replace('"','')
			ws.cell(row=rowno, column=colno).value = cell_value

	cur.close()
	conn.close()


def make_reference_page(ust_or_lust, wb):
	ws = wb.create_sheet('Reference')
	get_reference_data(ust_or_lust, ws)

	for row in ws[1:ws.max_row]:  
		cell = row[0]            
		cell.alignment = left_align
		cell.border = thin_border
		cell.font = Font(bold=True)
		cell = row[1]            
		cell.alignment = left_align
		cell.border = thin_border
		cell = row[2]            
		cell.alignment = center_align
		cell.border = thin_border
		cell = row[3]            
		cell.alignment = center_align
		cell.border = thin_border
		cell = row[4]            
		cell.alignment = center_align
		cell.border = thin_border
		cell = row[5]            
		cell.alignment = center_align
		cell.border = thin_border
		cell = row[6]            
		cell.alignment = left_align
		cell.border = thin_border
		cell = row[7]            
		cell.alignment = left_align
		cell.border = thin_border

	header_range = ws["A1:H1"]
	for row in header_range:
		for cell in row:
			cell.fill = fill_gen
			cell.alignment = center_align
			cell.font = Font(bold=True)


	ws.column_dimensions['A'].width = 69
	ws.column_dimensions['A'].font = Font(bold=True)
	ws.column_dimensions['B'].width = 30
	ws.column_dimensions['C'].width = 15
	ws.column_dimensions['D'].width = 8
	ws.column_dimensions['E'].width = 13
	ws.column_dimensions['F'].width = 13
	ws.column_dimensions['G'].width = 32
	ws.column_dimensions['H'].width = 70

	ws.freeze_panes = ws['A2']

	logger.info('Created reference page')


def make_lookup_page(organization_id, ust_or_lust, wb, lookup):
	pretty_name = lookup.replace('_',' ').title() 
	element_name = pretty_name.replace('_','')
	if pretty_name == 'Cause' or pretty_name == 'Source':
		pretty_name = pretty_name + 's'
	elif pretty_name == 'Corrective Action Strategy':
		pretty_name = 'Corrective Action Strategies'

	ws = wb.create_sheet(pretty_name)
	cell = ws.cell(row=1, column=1)
	cell.value = pretty_name
	cell.font = Font(bold=True)

	conn = utils.connect_db()
	cur = conn.cursor()	

	sql = f"select * from {lookup}"
	cur.execute(sql)
	data = cur.fetchall()

	for rowno, row in enumerate(data, start=2):
		for colno, cell_value in enumerate(row, start=1):
			ws.cell(row=rowno, column=colno).value = cell_value.replace('"','')

	sql = f"""select distinct state_value, epa_value
			from public.v_{ust_or_lust.lower()}_element_mapping 
			where organization_id = %s and element_name like '%%{element_name}%%'
			order by 1, 2"""
	cur.execute(sql, (organization_id, ))

	if cur.rowcount > 0:
		cell = ws.cell(row=1, column=3)
		cell.value = 'State Value'
		cell.font = Font(bold=True)

		cell = ws.cell(row=1, column=4)
		cell.value = 'EPA Value'
		cell.font = Font(bold=True)

		data = cur.fetchall()
		for rowno, row in enumerate(data, start=2):
			for colno, cell_value in enumerate(row, start=3):
				ws.cell(row=rowno, column=colno).value = cell_value		

	cur.close()
	conn.close()

	utils.autowidth(ws)

	logger.info('Created %s lookup page', lookup)


def make_substance_lookup_page(organization_id, ust_or_lust, wb):
	ws = wb.create_sheet('Substances')
	cell = ws.cell(row=1, column=1)
	cell.value = 'Substance Group'
	cell.font = Font(bold=True)
	cell.alignment = center_align

	cell = ws.cell(row=1, column=2)
	cell.value = 'Substance'
	cell.font = Font(bold=True)
	cell.alignment = center_align

	cell = ws.cell(row=1, column=3)
	cell.value = 'FederallyRegulated'
	cell.font = Font(bold=True)
	cell.alignment = center_align

	conn = utils.connect_db()
	cur = conn.cursor()	

	sql = f"select * from substances"
	cur.execute(sql)
	data = cur.fetchall()

	for rowno, row in enumerate(data, start=2):
		for colno, cell_value in enumerate(row, start=1):
			cell = ws.cell(row=rowno, column=colno)
			cell.value = cell_value.replace('"','')
			if colno == 3:
				cell.alignment = center_align
			if row[2] == 'N':
				cell.font = Font(bold=True)

	sql = f"""select distinct state_value, epa_value
			from public.v_{ust_or_lust.lower()}_element_mapping 
			where organization_id = %s and element_name like '%%Substance%%'
			order by 1, 2"""
	cur.execute(sql, (organization_id, ))

	if cur.rowcount > 0:
		cell = ws.cell(row=1, column=5)
		cell.value = 'State Value'
		cell.font = Font(bold=True)

		cell = ws.cell(row=1, column=6)
		cell.value = 'EPA Value'
		cell.font = Font(bold=True)

		data = cur.fetchall()
		for rowno, row in enumerate(data, start=2):
			for colno, cell_value in enumerate(row, start=5):
				ws.cell(row=rowno, column=colno).value = cell_value		

	cur.close()
	conn.close()

	utils.autowidth(ws)

	logger.info('Created substances lookup page')


def main(organization_id, ust_or_lust, data_only=False):
	logger.info('Exporting %s template for %s', ust_or_lust.upper(), organization_id.upper())

	file_path = config.local_ust_path + organization_id.upper() + '/'
	file_name = organization_id.upper() + '_' + ust_or_lust.upper() + '_template-' + str(date.today()) + '.xlsx'
	path = file_path + file_name
	
	wb = op.Workbook()
	make_data_page(organization_id, ust_or_lust, wb)
	if not data_only:
		make_reference_page(ust_or_lust, wb)

		make_lookup_page(organization_id, ust_or_lust, wb, 'facility_type')
		make_lookup_page(organization_id, ust_or_lust, wb, 'states')
		make_substance_lookup_page(organization_id, ust_or_lust, wb)
		if ust_or_lust.lower() == 'lust':
			make_lookup_page(organization_id, ust_or_lust, wb, 'source')
			make_lookup_page(organization_id, ust_or_lust, wb, 'cause')
			make_lookup_page(organization_id, ust_or_lust, wb, 'corrective_action_strategy')

	wb.save(path)
	
	logger.info('Exported %s to %s', file_name, file_path)


if __name__ == '__main__':   
	organization_id = 'MO'
	ust_or_lust = 'ust'
	main(organization_id, ust_or_lust)