from datetime import date
import ntpath
import os
from pathlib import Path
import sys  
ROOT_PATH = Path(__file__).parent.parent.parent
sys.path.append(os.path.join(ROOT_PATH, ''))

import openpyxl as op
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side

from python.state_processing import element_mapping_to_excel
from python.util import utils, config
from python.util.dataset import Dataset 
from python.util.logger_factory import logger


ust_or_release = 'ust' 			# Valid values are 'ust' or 'release'
control_id = 17                  # Enter an integer that is the ust_control_id or release_control_id

data_only = False				# Boolean; defaults to False. Set to True to generate a populated template that does not include the Reference and Lookup tabs.
template_only = False			# Boolean; defaults to False. Set to True to generate an blank template with no data.

# These variables can usually be left unset. This script will generate an Excel file in the appropriate state folder in the repo under /ust/python/exports/epa_templates.
# This file directory and its contents are excluded from pushes to the repo by .gitignore.
export_file_path = None
export_file_dir = None
export_file_name = None


gray_cell_fill = 'C9C9C9' # gray
green_cell_fill = '92D050' # green
yellow_cell_fill = 'FFFF00' # yellow
orange_cell_fill = 'FFC000' # orange
thin_border = Border(left=Side(style='thin'), 
					 right=Side(style='thin'), 
					 top=Side(style='thin'), 
					 bottom=Side(style='thin'))
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)


class Template:
	wb = None     

	def __init__(self, 
				 dataset,
				 data_only=False,
				 template_only=False):
		self.dataset = dataset
		self.data_only = data_only		
		self.template_only = template_only
		self.wb = op.Workbook()
		self.wb.save(self.dataset.export_file_path)
		if not self.data_only:
			self.make_reference_tab()
			self.make_lookup_tabs()
			if not self.template_only:
				element_mapping_to_excel.build_ws(self.dataset, self.wb.create_sheet(), admin=False)
				self.make_mapping_tabs()
		self.make_data_tabs()	
		self.cleanup_wb()
		logger.info('Template exported to %s', self.dataset.export_file_path)


	def cleanup_wb(self):
		try:
			self.wb.remove(self.wb['Sheet'])
		except:
			pass
		self.wb.save(self.dataset.export_file_path)


	def make_reference_tab(self):
		ws = self.wb.create_sheet('Reference')

		headers = utils.get_headers(f'v_{self.dataset.ust_or_release}_elements')
		for colno, header in enumerate(headers, start=1):
			ws.cell(row=1, column=colno).value = header
		conn = utils.connect_db()
		cur = conn.cursor()
		sql = f"select * from public.v_{self.dataset.ust_or_release}_elements"	
		cur.execute(sql)
		data = cur.fetchall()
		for rowno, row in enumerate(data, start=2):
			for colno, cell_value in enumerate(row, start=1):
				if (colno == 8 or colno == 7) and cell_value:
					cell_value = cell_value.replace('"','')
				ws.cell(row=rowno, column=colno).value = cell_value
		cur.close()
		conn.close()

		if ust_or_release == 'ust':
			for row in ws[1:ws.max_row]:  
				cell = row[0]            
				cell.alignment = left_align
				cell.border = thin_border
				cell = row[1]            
				cell.alignment = left_align
				cell.border = thin_border
				cell.font = Font(bold=True)
				cell = row[2]            
				cell.alignment = left_align
				cell.border = thin_border
				cell = row[3]            
				cell.alignment = center_align
				cell.border = thin_border
				cell = row[4]            
				cell.alignment = center_align
				cell.border = thin_border
				cell = row[5]            
				cell.alignment = center_align
				cell.border = thin_border
				cell = row[6]            
				cell.alignment = center_align
				cell.border = thin_border
				cell = row[7]            
				cell.alignment = left_align
				cell.border = thin_border
				cell = row[8]            
				cell.alignment = left_align
				cell.border = thin_border
			header_range = ws["A1:I1"]
			for row in header_range:
				for cell in row:
					cell.fill = utils.get_fill_gen(gray_cell_fill)
					cell.alignment = center_align
					cell.font = Font(bold=True)
			element_name_range = ws["B2:B200"]
			green_cells = ['FacilityID','TankID','CompartmentID','PipingID']
			for row in element_name_range:
			    for cell in row:
			        if cell.value in green_cells:
			        	cell.fill = utils.get_fill_gen(green_cell_fill)
			        	if cell.value != 'FacilityID':
				        	cell.offset(row=0, column=1).fill = utils.get_fill_gen(yellow_cell_fill)
			ws.column_dimensions['A'].width = 14
			ws.column_dimensions['B'].width = 69
			ws.column_dimensions['B'].font = Font(bold=True)
			ws.column_dimensions['C'].width = 30
			ws.column_dimensions['D'].width = 15
			ws.column_dimensions['E'].width = 8
			ws.column_dimensions['F'].width = 13
			ws.column_dimensions['G'].width = 13
			ws.column_dimensions['H'].width = 32
			ws.column_dimensions['I'].width = 70
		elif ust_or_release == 'release':
			for row in ws[1:ws.max_row]:  
				cell = row[0]            
				cell.alignment = left_align
				cell.border = thin_border
				cell.font = Font(bold=True)
				cell = row[1]            
				cell.alignment = left_align
				cell.border = thin_border
				cell = row[2]            
				cell.alignment = left_align
				cell.border = thin_border
				cell = row[3]            
				cell.alignment = center_align
				cell.border = thin_border
				cell = row[4]            
				cell.alignment = center_align
				cell.border = thin_border
				cell = row[5]            
				cell.alignment = center_align
				cell.border = thin_border
				cell = row[6]            
				cell.alignment = left_align
				cell.border = thin_border
				cell = row[7]            
				cell.alignment = left_align
				cell.border = thin_border
			header_range = ws["A1:H1"]
			for row in header_range:
				for cell in row:
					cell.fill = utils.get_fill_gen(gray_cell_fill)
					cell.alignment = center_align
					cell.font = Font(bold=True)
			ws.column_dimensions['A'].width = 69
			ws.column_dimensions['A'].font = Font(bold=True)
			ws.column_dimensions['B'].width = 30
			ws.column_dimensions['C'].width = 15
			ws.column_dimensions['D'].width = 8
			ws.column_dimensions['E'].width = 13
			ws.column_dimensions['F'].width = 13
			ws.column_dimensions['G'].width = 32
			ws.column_dimensions['H'].width = 70

		ws.freeze_panes = ws['A2']
		self.wb.save(self.dataset.export_file_path)
		logger.info('Created Reference tab')


	def make_lookup_tabs(self):
		lookups = utils.get_lookup_tabs(ust_or_release=self.dataset.ust_or_release)
		for lookup in lookups:
			self.make_lookup_tab(lookup)
		self.wb.save(self.dataset.export_file_path)


	def make_lookup_tab(self, lookup):
		lookup_table_name = lookup[0]
		lookup_column_name = lookup[1]
		lookup_column_id = lookup[2]
		logger.info('Working on lookup table %s, column %s', lookup_table_name, lookup_column_name)
		if lookup_table_name == 'substances':
			ws = self.wb.create_sheet('Substances lookup')
			self.make_substance_lookup_tab(ws)
			logger.info('Created Substances lookup tab')
			return

		if lookup_table_name == 'corrective_action_strategies':
			pretty_name = 'Corrective Actions'
		else:
			pretty_name = lookup_table_name.replace('_',' ').title()
		ws = self.wb.create_sheet(pretty_name + ' lookup')
		
		conn = utils.connect_db()
		cur = conn.cursor()	
		
		sql = f"select {lookup_column_name} from public.{lookup_table_name} order by 1"
		cur.execute(sql)
		data = cur.fetchall()
		for rowno, row in enumerate(data, start=2):
			for colno, cell_value in enumerate(row, start=1):
				ws.cell(row=rowno, column=colno).value = cell_value.replace('"','')
		cell = ws.cell(row=1, column=1)
		cell.value = pretty_name
		cell.font = Font(bold=True)
		utils.autowidth(ws)

		cur.close()
		conn.close()
		logger.info('Created %s lookup tab', pretty_name)


	def make_substance_lookup_tab(self, ws):
		cell = ws.cell(row=1, column=1)
		cell.value = 'Substance Group'
		cell.font = Font(bold=True)
		cell.alignment = center_align

		cell = ws.cell(row=1, column=2)
		cell.value = 'Substance'
		cell.font = Font(bold=True)
		cell.alignment = center_align

		conn = utils.connect_db()
		cur = conn.cursor()	
		sql = f"select substance_group, substance from public.substances order by 1, 2"
		cur.execute(sql)
		data = cur.fetchall()
		for rowno, row in enumerate(data, start=2):
			for colno, cell_value in enumerate(row, start=1):
				cell = ws.cell(row=rowno, column=colno)
				cell.value = cell_value.replace('"','')
		utils.autowidth(ws)
		cur.close()
		conn.close()


	def get_mapping_tabs(self):
		conn = utils.connect_db()
		cur = conn.cursor()	
		sql = f"""select epa_table_name, epa_column_name, database_lookup_table, database_lookup_column   
				from public.v_{self.dataset.ust_or_release}_available_mapping
				where {self.dataset.ust_or_release}_control_id = %s order by 1, 2"""
		cur.execute(sql, (self.dataset.control_id,))
		rows = cur.fetchall()
		cur.close()
		conn.close()
		return rows


	def make_mapping_tabs(self):
		mappings = self.get_mapping_tabs()
		for mapping in mappings:
			self.make_mapping_tab(mapping)
		self.wb.save(self.dataset.export_file_path)


	def make_mapping_tab(self, mapping):
		mapping_table_name = mapping[0]
		mapping_column_name = mapping[1]
		database_lookup_table = mapping[2]
		database_lookup_column = mapping[3]
		logger.info('Working on mapping table %s, column %s using lookup table %s, column %s',
			 mapping_table_name, mapping_column_name, database_lookup_table, database_lookup_column)
		if database_lookup_table == 'substances':
			self.make_substance_mapping_tab()
			logger.info('Created Substances mapping tab')
			return
		pretty_name = database_lookup_table.replace('_',' ').title()
		# avoid tab names longer than 31 characters
		if pretty_name == 'Tank Material Descriptions':
			pretty_name = 'Tank Material Desc'
		elif pretty_name == 'Tank Secondary Containments':
			pretty_name = 'Secondary Containment'
		elif pretty_name == 'Pipe Tank Top Sump Wall Types':
			pretty_name = 'Pipe Top Sump Wall Type'
		elif pretty_name == 'Corrective Action Strategies':
			pretty_name = 'Corrective Actions'

		ws = self.wb.create_sheet(pretty_name + ' mapping')

		cell = ws.cell(row=1, column=1)
		cell.value = pretty_name
		cell.font = Font(bold=True)

		conn = utils.connect_db()
		cur = conn.cursor()	

		sql = f"select {database_lookup_column} from {database_lookup_table} order by 1"
		cur.execute(sql)
		data = cur.fetchall()
		for rowno, row in enumerate(data, start=2):
			for colno, cell_value in enumerate(row, start=1):
				ws.cell(row=rowno, column=colno).value = cell_value.replace('"','')

		sql = f"""select distinct organization_value, epa_value, programmer_comments, epa_comments, organization_comments
				from public.v_{self.dataset.ust_or_release}_element_mapping 
				where {self.dataset.ust_or_release}_control_id = %s and epa_column_name = %s
				order by 1, 2"""
		cur.execute(sql, (self.dataset.control_id, mapping_column_name))

		if cur.rowcount > 0:
			cell = ws.cell(row=1, column=3)
			cell.value = 'Organization Value'
			cell.font = Font(bold=True)
			cell = ws.cell(row=1, column=4)
			cell.value = 'EPA Value'
			cell.font = Font(bold=True)
			cell = ws.cell(row=1, column=5)
			cell.value = 'Programmer Comments'
			cell.font = Font(bold=True)
			cell = ws.cell(row=1, column=6)
			cell.value = 'EPA Comments'
			cell.font = Font(bold=True)
			cell = ws.cell(row=1, column=7)
			cell.value = 'Organization Comments'
			cell.font = Font(bold=True)

			data = cur.fetchall()
			for rowno, row in enumerate(data, start=2):
				for colno, cell_value in enumerate(row, start=3):
					ws.cell(row=rowno, column=colno).value = cell_value		

		utils.autowidth(ws)

		cur.close()
		conn.close()

		logger.info('Created %s mapping tab', pretty_name)


	def make_substance_mapping_tab(self):
		ws = self.wb.create_sheet('Substances mapping')
		self.make_substance_lookup_tab(ws)

		conn = utils.connect_db()
		cur = conn.cursor()	
		sql = f"""select distinct organization_value, epa_value, programmer_comments, epa_comments, organization_comments
				from public.v_{self.dataset.ust_or_release}_element_mapping 
				where {self.dataset.ust_or_release}_control_id = %s and epa_column_name = 'substance_id'
				order by 1, 2"""
		cur.execute(sql, (self.dataset.control_id,))

		if cur.rowcount > 0:
			cell = ws.cell(row=1, column=4)
			cell.value = 'Organization Value'
			cell.font = Font(bold=True)
			cell = ws.cell(row=1, column=5)
			cell.value = 'EPA Value'
			cell.font = Font(bold=True)
			cell = ws.cell(row=1, column=6)
			cell.value = 'Programmer Comments'
			cell.font = Font(bold=True)
			cell = ws.cell(row=1, column=7)
			cell.value = 'EPA Comments'
			cell.font = Font(bold=True)
			cell = ws.cell(row=1, column=8)
			cell.value = 'Organization Comments'
			cell.font = Font(bold=True)
			data = cur.fetchall()
			for rowno, row in enumerate(data, start=2):
				for colno, cell_value in enumerate(row, start=4):
					ws.cell(row=rowno, column=colno).value = cell_value		
		utils.autowidth(ws)

		cur.close()
		conn.close()
		logger.info('Created Substances mapping tab')


	def make_data_tabs(self):
		tabs = utils.get_data_tabs(ust_or_release=self.dataset.ust_or_release)
		for tab in tabs:
			self.make_data_tab(tab)
		self.wb.save(self.dataset.export_file_path)


	def make_data_tab(self, tab):
		view_name = tab[0]
		tab_name = tab[1]
		ws = self.wb.create_sheet(tab_name)
		ws.title = tab_name
		headers = utils.get_headers(view_name)
		green_cells = []
		orange_cells = []
		if self.dataset.ust_or_release == 'ust':
			if tab_name == 'Facility':
				green_cells = ['FacilityID']
			elif tab_name == 'Facility Dispenser':
				green_cells = ['DispenserID']
				orange_cells = ['FacilityID']
			elif tab_name == 'Tank':
				green_cells = ['TankID']
				orange_cells = ['FacilityID']
			elif tab_name == 'Tank Substance' or tab_name == 'Tank Dispenser':
				green_cells = ['Substance','DispenserID']
				orange_cells = ['FacilityID','TankID','TankName']
			elif tab_name == 'Compartment':
				green_cells = ['CompartmentID']
				orange_cells = ['FacilityID','TankID','TankName']
			elif tab_name == 'Piping' or tab_name == 'Compartment Substance' or tab_name == 'Compartment Dispenser':
				green_cells = ['Substance','PipingID','DispenserID']
				orange_cells = ['FacilityID','TankID','TankName','CompartmentID','CompartmentName']
		for colno, header in enumerate(headers, start=1):
			cell = ws.cell(row=1, column=colno)
			cell.value = header
			cell.font = Font(bold=True)
			if cell.value in green_cells:
				cell.fill = utils.get_fill_gen(green_cell_fill)
			if cell.value in orange_cells:
				cell.fill = utils.get_fill_gen(orange_cell_fill)
		if not self.template_only:
			conn = utils.connect_db()
			cur = conn.cursor()
			sql = f"select * from public.{view_name} where {self.dataset.ust_or_release}_control_id = %s"
			cur.execute(sql, (self.dataset.control_id,))
			data = cur.fetchall()
			for rowno, row in enumerate(data, start=2):
				for colno, cell_value in enumerate(row, start=1):
					ws.cell(row=rowno, column=colno).value = cell_value
			cur.close()
			conn.close()
		ws.delete_cols(1)
		utils.autowidth(ws)
		ws.freeze_panes = ws['A2']
		logger.info('Created %s tab', tab_name)



def main(ust_or_release, control_id=None, data_only=False, template_only=False, export_file_name=None, export_file_dir=None, export_file_path=None):
	if template_only and control_id == 0:
		control_id = 1

	dataset = Dataset(ust_or_release=ust_or_release,
				 	  control_id=control_id, 
				 	  base_file_name='template_' + utils.get_timestamp_str() + '.xlsx',
					  export_file_name=export_file_name,
					  export_file_dir=export_file_dir,
					  export_file_path=export_file_path)

	template = Template(dataset=dataset,
						data_only=data_only,
						template_only=template_only)


if __name__ == '__main__':   
	main(ust_or_release=ust_or_release,
		 control_id=control_id, 
		 data_only=data_only, 
		 template_only=template_only,
		 export_file_name=export_file_name,
		 export_file_dir=export_file_dir,
		 export_file_path=export_file_path)
