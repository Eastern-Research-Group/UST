import os
from pathlib import Path
import sys  
ROOT_PATH = Path(__file__).parent.parent.parent
sys.path.append(os.path.join(ROOT_PATH, ''))

import openpyxl as op
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side

from python.util.logger_factory import logger
from python.util import utils


ust_or_release = 'release' # valid values are 'ust' or 'release'
control_id = 6

yellow_cell_fill = 'FFFF00' # yellow
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)


class PopulatedTemplate():
	def __init__(self, ust_or_release, control_id):
		self.ust_or_release = ust_or_release.lower()
		if self.ust_or_release not in ['ust','release']:
			logger.error('Invalid value %s for ust_or_release. Valid values are ust or release. Exiting...', ust_or_release)
			exit()
		self.control_id = control_id  
		self.schema = utils.get_schema_from_control_id(control_id, ust_or_release)
		self.organization_id = utils.get_org_from_control_id(control_id, ust_or_release)
		self.export_file_name = None 
		self.export_file_path = None  
		self.populate_export_vars()
		self.wb = None


	def populate_export_vars(self):
		if self.ust_or_release == 'ust':
			uor = 'UST'
		elif self.ust_or_release == 'release':
			uor = 'Releases'		
		self.export_file_name = self.organization_id.upper() + '_' + uor + '_mapping_' + utils.get_timestamp_str() + '.xlsx'
		self.export_file_dir = '../exports/mapping/' + self.organization_id.upper() + '/'
		Path(self.export_file_dir).mkdir(parents=True, exist_ok=True)
		self.export_file_path = self.export_file_dir + self.export_file_name


	def populate_element_mapping(self):
		conn = utils.connect_db()
		cur = conn.cursor()
		sql = f"""select 'insert into {self.ust_or_release}_element_mapping ({self.ust_or_release}_control_id, epa_table_name, epa_column_name, organization_table_name, organization_column_name, programmer_comments) values ({self.control_id}, ''' ||
					epa_table_name || ''',''' || epa_column_name || ''',''' || org_table_name || ''',''' || org_column_name || ''',''Direct mapping to EPA template per state'')'
				from 
					(select distinct c.table_name as epa_table_name, b.database_column_name as epa_column_name, 
						a.table_name as org_table_name, a.column_name as org_column_name, a.ordinal_position
					from information_schema.columns a join {self.ust_or_release}_elements b on a.column_name = b.element_name 
						join {self.ust_or_release}_elements_tables c on b.element_id = c.element_id 
					where table_schema = '{self.schema}') x 
				where not exists
					(select 1 from {self.ust_or_release}_element_mapping d 
					where d.{self.ust_or_release}_control_id = {self.control_id} and x.org_table_name = d.organization_table_name and x.org_column_name = d.organization_column_name)
				order by org_table_name, ordinal_position"""
		print(sql)
		exit()
		cur.execute(sql, (self.schema, ))
		rows = cur.fetchall()
		i = 0
		for row in rows:
			cur.execute(row[0])
			i += 1
		conn.commit()

		if self.ust_or_release == 'release':
			sql = f"""select 'select distinct "' || column_name || '"::text as cause from {self.schema}."' || table_name || '"  where "' || column_name || '" is not null union all '
					from information_schema.columns 
					where table_schema = %s
					and column_name like 'CauseOfRelease%%'
					order by 1"""
			cur.execute(sql, (self.schema,))
			rows = cur.fetchall()
			if rows:
				sql = f"create or replace view {self.schema}.v_unique_causes as select distinct cause from ("
				subquery = ''
				for row in rows:
					subquery = subquery + row[0] 
				subquery = subquery[:-10]
				sql = sql + subquery + ') x'
				cur.execute(sql)
				logger.info('Created view %s.v_unique_causes', self.schema)
				self.insert_column_mapping('v_unique_causes', 'cause', 'ust_release_cause','cause_id');
				i += 1

			sql = f"""select 'select distinct "' || column_name || '"::text as source from {self.schema}."' || table_name || '"  where "' || column_name || '" is not null union all '
					from information_schema.columns 
					where table_schema = %s
					and column_name like 'SourceOfRelease%%'
					order by 1"""
			cur.execute(sql, (self.schema,))
			rows = cur.fetchall()
			if rows:
				sql = f"create or replace view {self.schema}.v_unique_sources as select distinct source from ("
				subquery = ''
				for row in rows:
					subquery = subquery + row[0] 
				subquery = subquery[:-10]
				sql = sql + subquery + ') x'
				cur.execute(sql)
				logger.info('Created view %s.v_unique_sources', self.schema)
				self.insert_column_mapping('v_unique_sources', 'source', 'ust_release_source','source_id');
				i += 1

			sql = f"""select 'select distinct "' || column_name || '"::text as substance from {self.schema}."' || table_name || '"  where "' || column_name || '" is not null union all '
					from information_schema.columns 
					where table_schema = %s
					and column_name like 'SubstanceReleased%%'
					order by 1"""
			cur.execute(sql, (self.schema,))
			rows = cur.fetchall()
			if rows:
				sql = f"create or replace view {self.schema}.v_unique_substances as select distinct substance from ("
				subquery = ''
				for row in rows:
					subquery = subquery + row[0] 
				subquery = subquery[:-10]
				sql = sql + subquery + ') x'
				cur.execute(sql)
				logger.info('Created view %s.v_unique_substances', self.schema)
				self.insert_column_mapping('v_unique_substances', 'substance', 'ust_release_substance','substance_id');
				i += 1

			sql = f"""select 'select distinct "' || column_name || '"::text as corrective_action_strategy from {self.schema}."' || table_name || '"  where "' || column_name || '" is not null union all '
					from information_schema.columns 
					where table_schema = %s
					and column_name like 'CorrectiveActionStrategy%%' and column_name not like '%%StartDate'
					order by 1"""
			cur.execute(sql, (self.schema,))
			rows = cur.fetchall()
			if rows:
				sql = f"create or replace view {self.schema}.v_unique_corrective_action_strategies as select distinct corrective_action_strategy from ("
				subquery = ''
				for row in rows:
					subquery = subquery + row[0] 
				subquery = subquery[:-10]
				sql = sql + subquery + ') x'
				cur.execute(sql)
				logger.info('Created view %s.v_unique_corrective_action_strategies', self.schema)
				self.insert_column_mapping('v_unique_corrective_action_strategies', 'corrective_action_strategy', 'ust_release_corrective_action_strategy','corrective_action_strategy_id');
				i += 1

			sql = f"""select table_name, column_name from information_schema.columns a
			          where table_schema = %s and column_name = 'LUSTID'
			          and not exists 
			          	(select 1 from release_element_mapping b 
			          	where release_control_id = %s
			          	and a.table_name = b.organization_table_name and a.column_name = b.organization_column_name)"""
			cur.execute(sql, (self.schema, self.control_id))
			rows = cur.fetchall()
			for row in rows:
				self.insert_column_mapping(row[0], row[1],'ust_release','release_id');
				i += 1

			sql = f"""select table_name, column_name from information_schema.columns a
			          where table_schema = %s and column_name = 'LUSTStatus'
			          and not exists 
			          	(select 1 from release_element_mapping b 
			          	where release_control_id = %s
			          	and a.table_name = b.organization_table_name and a.column_name = b.organization_column_name)"""
			cur.execute(sql, (self.schema, self.control_id))
			rows = cur.fetchall()
			for row in rows:
				self.insert_column_mapping(row[0], row[1],'ust_release','release_status_id');
				i += 1

			sql = f"""select table_name, column_name from information_schema.columns a
			          where table_schema = %s and 
			          (column_name like 'Unit%%' or column_name like 'QuantityReleased%%')
			          and not exists 
			          	(select 1 from release_element_mapping b 
			          	where release_control_id = %s
			          	and a.table_name = b.organization_table_name and a.column_name = b.organization_column_name)"""
			cur.execute(sql, (self.schema, self.control_id))
			rows = cur.fetchall()
			for row in rows:
				table_name = row[0]
				column_name = row[1]
				if 'Substance' in column_name:
					epa_column_name = 'substance_id'
				elif 'QuantityReleased' in column_name:
					epa_column_name = 'quantity_release'
				elif 'Unit' in column_name:
					epa_column_name = 'unit'
				self.insert_column_mapping(table_name, column_name, 'ust_release_substance', epa_column_name);
				i += 1

			sql = f"""select table_name, column_name from information_schema.columns a
			          where table_schema = %s and column_name like 'CorrectiveActionStartDate%%' 
			          and not exists 
			          	(select 1 from release_element_mapping b 
			          	where release_control_id = %s
			          	and a.table_name = b.organization_table_name and a.column_name = b.organization_column_name)"""
			cur.execute(sql, (self.schema, self.control_id))
			rows = cur.fetchall()
			for row in rows:
				self.insert_column_mapping(row[0], row[1], 'ust_release_corrective_action_strategy', 'corrective_action_strategy_start_date');
				i += 1

		logger.info('Inserted %s rows into %s_element_mapping', i, self.ust_or_release)
		cur.close()
		conn.close()	
		self.check_for_unmapped_elements()


	def check_for_unmapped_elements(self):
		conn = utils.connect_db()
		cur = conn.cursor()
		sql = f"""select b.table_name, a.database_column_name 
				from {self.ust_or_release}_elements a join {self.ust_or_release}_elements_tables b on a.element_id = b.element_id 
				where not exists 
					(select 1 from {self.ust_or_release}_element_mapping c
					where {self.ust_or_release}_control_id = %s
					and c.epa_table_name = b.table_name and c.epa_column_name = a.database_column_name)
				and exists 
					(select 1 from information_schema.columns d
					where table_schema = %s
					and a.element_name = d.column_name)
				order by 1, 2"""
		cur.execute(sql, (self.control_id, self.schema))
		rows = cur.fetchall()
		all_mapped = True
		for row in rows:
			all_mapped = False
			logger.warning('Unmapped state column %s.%s', row[0], row[1])
		cur.close()
		conn.close()	
		if not all_mapped:
			logger.warning('Please resolved the unmapped state columns and re-run this script. You can use the insert_column_mapping() function of this script.')
			exit()


	def insert_column_mapping(self, org_table_name, org_column_name, epa_table_name, epa_column_name, programmer_comments=None):
		conn = utils.connect_db()
		cur = conn.cursor()
		sql = f"""select count(*) from {self.ust_or_release}_element_mapping 
				where {self.ust_or_release}_control_id = %s 
				and epa_table_name = %s and epa_column_name = %s and organization_table_name = %s and organization_column_name = %s"""
		cur.execute(sql, (self.control_id, epa_table_name, epa_column_name, org_table_name, org_column_name))
		if cur.fetchone()[0] == 0:
			if programmer_comments:
				sql = f"""insert into {self.ust_or_release}_element_mapping ({self.ust_or_release}_control_id, epa_table_name, epa_column_name, organization_table_name, organization_column_name, programmer_comments) 
						  values (%s, %s, %s, %s, %s, %s)"""
				cur.execute(sql, (self.control_id, epa_table_name, epa_column_name, org_table_name, org_column_name, programmer_comments))
			else:
				sql = f"""insert into {self.ust_or_release}_element_mapping ({self.ust_or_release}_control_id, epa_table_name, epa_column_name, organization_table_name, organization_column_name) 
						  values (%s, %s, %s, %s, %s)"""
				cur.execute(sql, (self.control_id, epa_table_name, epa_column_name, org_table_name, org_column_name))
			conn.commit()
			logger.info('Inserted mapping from %s.%s to %s.%s', org_table_name, org_column_name, epa_table_name, epa_column_name)
		cur.close()
		conn.close()	


	def cleanup_wb(self):
		try:
			self.wb.remove(self.wb['Sheet'])
		except:
			pass
		self.wb.save(self.export_file_path)


	def populate_lookup_values(self, ws, db_lookup_table, db_lookup_column):
		if db_lookup_column == 'substance':
			cell = ws.cell(row=1, column=1)
			cell.value = 'Substance Group'
			cell.font = Font(bold=True)
			cell = ws.cell(row=1, column=2)
			cell.value = 'Substance'
			cell.font = Font(bold=True)
		else:
			cell = ws.cell(row=1, column=1)
			cell.value = db_lookup_column
		
		cell.font = Font(bold=True)
		cell = ws.cell(row=1, column=4)
		cell.value = 'Organization Value'
		cell.font = Font(bold=True)
		cell = ws.cell(row=1, column=5)
		cell.value = 'EPA Value'
		cell.font = Font(bold=True)
		cell = ws.cell(row=1, column=6)
		cell.value = 'Programmer Comments'
		cell.font = Font(bold=True)
		cell = ws.cell(row=1, column=7)
		cell.value = 'EPA Comments'
		cell.font = Font(bold=True)
		cell = ws.cell(row=1, column=8)
		cell.value = 'Organization Comments'
		cell.font = Font(bold=True)
		cell = ws.cell(row=1, column=9)
		cell.value = 'EPA Approved'
		cell.font = Font(bold=True)
		cell = ws.cell(row=1, column=10)
		cell.value = 'Exclude from query'
		cell.font = Font(bold=True)

		if db_lookup_column == 'substance':
			conn = utils.connect_db()
			cur = conn.cursor()	
			sql = f"select substance_group, substance from public.substances order by 1, 2"
			cur.execute(sql)
			data = cur.fetchall()
			for rowno, row in enumerate(data, start=2):
				for colno, cell_value in enumerate(row, start=1):
					cell = ws.cell(row=rowno, column=colno)
					cell.value = cell_value.replace('"','')
			utils.autowidth(ws)
			cur.close()
			conn.close()
		else:
			epa_values = utils.get_table_values(db_lookup_table, db_lookup_column)
			i = 2
			for val in epa_values:
				cell = ws.cell(row=i, column=1)
				cell.value = val
				i += 1


	def check_state_mapping(self):
		logger.info('Working on State Mapping...')

		self.wb = op.Workbook()

		conn = utils.connect_db()
		cur = conn.cursor()	

		sql = f"""select distinct epa_table_name, epa_column_name, {self.ust_or_release}_element_mapping_id, organization_table_name, organization_column_name, 
						database_lookup_table, database_lookup_column, table_sort_order, column_sort_order
				from v_{self.ust_or_release}_needed_mapping 
				where {self.ust_or_release}_control_id = %s order by table_sort_order, column_sort_order"""
		cur.execute(sql, (self.control_id,))
		rows = cur.fetchall()
		for row in rows:
			epa_table_name = row[0]
			epa_column_name = row[1]
			element_mapping_id = row[2]
			org_table_name = row[3]
			org_column_name = row[4]
			db_lookup_table = row[5]
			db_lookup_column = row[6]
			
			logger.info('Working on EPA column %s.%s', epa_table_name, epa_column_name)
			
			sql2 = f"""select distinct a."{org_column_name}", b.{db_lookup_column} 
						from {self.schema}."{org_table_name}" a left join public.{db_lookup_table} b 
						on a."{org_column_name}"::text  = b.{db_lookup_column} 
						where a."{org_column_name}" is not null 
						order by 1"""
			cur.execute(sql2)
			if cur.rowcount > 0:
				ws = self.wb.create_sheet(org_column_name)
				self.populate_lookup_values(ws, db_lookup_table, db_lookup_column)
				data = cur.fetchall()
				for rowno, row in enumerate(data, start=2):
					for colno, cell_value in enumerate(row, start=4):
						cell = ws.cell(row=rowno, column=colno)
						cell.value = cell_value		
						if not cell_value:
							cell.fill = utils.get_fill_gen(yellow_cell_fill)
				utils.autowidth(ws)
				ws.cell(row=1, column=13).value = self.ust_or_release + '_element_mapping_id'
				ws.cell(row=2, column=13).value = element_mapping_id

			self.cleanup_wb()
			self.wb.save(self.export_file_path)

		cur.close()
		conn.close()	


	def upload_mapping(self, ws):
		conn = utils.connect_db()
		cur = conn.cursor()
		element_mapping_id = ws['L2'].value
		logger.info('element_mapping_id = %s', element_mapping_id)
		r = 1
		mapping_found = True
		while mapping_found:
			r += 1
			organization_value = ws.cell(row=r, column=4).value 
			if not organization_value:
				mapping_found = False 
				continue
			epa_value = ws.cell(row=r, column=5).value
			programmer_comments = ws.cell(row=r, column=6).value
			epa_comments = ws.cell(row=r, column=7).value
			organization_comments = ws.cell(row=r, column=8).value
			epa_approved = ws.cell(row=r, column=9).value
			exclude_from_query = ws.cell(row=r, column=10).value

			sql = f"""insert into {self.ust_or_release}_element_value_mapping ({self.ust_or_release}_element_mapping_id, organization_value, epa_value, programmer_comments, epa_comments, organization_comments, epa_approved, exclude_from_query)
			          values(%s, %s, %s, %s, %s, %s, %s, %s) 
			          on conflict ({self.ust_or_release}_element_mapping_id, organization_value, epa_value)
			          do update set (organization_value, epa_value, programmer_comments, epa_comments, organization_comments, epa_approved, exclude_from_query) = 
			           (excluded.organization_value, excluded.epa_value, excluded.programmer_comments, excluded.epa_comments, excluded.organization_comments, excluded.epa_approved, excluded.exclude_from_query)"""
			cur.execute(sql, (element_mapping_id, organization_value, epa_value, programmer_comments, epa_comments, organization_comments, epa_approved, exclude_from_query))
			conn.commit()
			logger.info("Inserted organization value '%s' = EPA value '%s'", organization_value, epa_value)
		cur.close()
		conn.close()


	def insert_element_value_mapping(self, mapping_file_path, sheetname=None):
		wb = op.load_workbook(mapping_file_path)
		if sheetname:
			logger.info('Working on %s', sheetname)
			ws = wb[sheetname]
			# check if this sheet is a mapping page
			if ws['D1'].value == 'Organization Value':
				self.upload_mapping(ws)
			else:
				logger.info('Worksheet %s does not appear to be a mapping tab; ignoring', sheetname)
		else: # go through all worksheets in workbook
			for sheetname in wb.sheetnames:
				logger.info('Working on %s', sheetname)
				ws = wb[sheetname]
				if ws['D1'].value == 'Organization Value':
					self.upload_mapping(ws)
			

	def check_missing_mapping(self):
		conn = utils.connect_db()
		cur = conn.cursor()
		sql = f"""select organization_table_name, organization_column_name
				from 
					(select distinct organization_table_name, organization_column_name, epa_table_name, epa_column_name, table_sort_order, column_sort_order
					from v_{self.ust_or_release}_needed_mapping 
					where {self.ust_or_release}_control_id = %s and mapping_complete = 'N'
					order by table_sort_order, column_sort_order) x"""
		cur.execute(sql, (self.control_id,))
		rows = cur.fetchall()
		for row in rows:
			org_table_name = row[0]
			org_column_name = row[1]
			sql2 = f"""select count(*) from {self.schema}."{org_table_name}" where "{org_column_name}" is not null"""
			cur.execute(sql2)
			cnt = cur.fetchone()[0]
			if cnt > 0:
				logger.warning('No value mapping for %s.%s!', org_table_name, org_column_name) 
		cur.close()
		conn.close()


	def check_incomplete_mapping(self):
		conn = utils.connect_db()
		cur = conn.cursor()



		cur.close()
		conn.close()

if __name__ == '__main__':       
	pop_temp = PopulatedTemplate(ust_or_release, control_id)

	# # Step 1: populate mapping
	pop_temp.populate_element_mapping()
	# # If necessary, manually insert missing mapping (the function above will print a list of unmapped elements)
	# pop_temp.insert_column_mapping('ust_facility', 'AssociatedLUSTID', 'ust_facility', 'AssociatedUSTReleaseID')

	# pop_temp.insert_column_mapping('ust_release','CorrectiveActionStrategy1StartDate','ust_release_corrective_action_strategy','corrective_action_strategy_start_date');
	# pop_temp.insert_column_mapping('ust_release','CorrectiveActionStrategy2StartDate','ust_release_corrective_action_strategy','corrective_action_strategy_start_date');
	# pop_temp.insert_column_mapping('ust_release','CorrectiveActionStrategy3StartDate','ust_release_corrective_action_strategy','corrective_action_strategy_start_date');

	# # Export a workbork pre-populated with exact value mapping matches 
	# pop_temp.check_state_mapping()

	# # Step 2: Complete any missing mapping in workbook exported above and insert all mappings into database 
	mapping_file_path = fr'C:\Users\erguser\repos\ERG\UST\ust\python\exports\mapping\AZ\AZ_Releases_mapping_20240724192944.xlsx'
	# pop_temp.insert_element_value_mapping(mapping_file_path)
	
	# # Step 3: Check that all mapping is complete 
	# pop_temp.check_missing_mapping()
	# pop_temp.check_incomplete_mapping()

	# # Step 4: After receiving state/EPA feedback, you can either re-run Step 2 to re-insert the entire workbook into the database, or process individual mapping tabs:
	pop_temp.insert_element_value_mapping(mapping_file_path, sheetname='corrective_action_strategy')
