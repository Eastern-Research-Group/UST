from datetime import date
import ntpath
import os
from pathlib import Path
import sys  
ROOT_PATH = Path(__file__).parent.parent.parent
sys.path.append(os.path.join(ROOT_PATH, ''))

import openpyxl as op
from openpyxl.styles import Alignment, Font

from python.util.logger_factory import logger
from python.util import utils
from python.util.dataset import Dataset 

ust_or_release = 'ust' 			# Valid values are 'ust' or 'release'
control_id = 0                  # Enter an integer that is the ust_control_id or release_control_id

# These variables can usually be left unset. This script will general an Excel file in the appropriate state folder in the repo under /ust/python/exports/control_table_summaries
# This file directory and its contents are excluded from pushes to the repo by .gitignore.
export_file_path = None
export_file_dir = None
export_file_name = None


yellow_cell_fill = 'FFFF00' # yellow

class Summary:
	wb = None     

	def __init__(self, 
				 dataset):
		self.dataset = dataset
		self.wb = op.Workbook()
		self.wb.save(self.dataset.export_file_path)
		self.export_summary()
		self.export_row_counts()
		self.performance_measure_comparison()
		self.mapped_element_summary()
		self.cleanup_wb()
		logger.info('Summary exported to %s', self.dataset.export_file_path)


	def cleanup_wb(self):
		try:
			self.wb.remove(self.wb['Sheet'])
		except:
			pass
		self.wb.save(self.dataset.export_file_path)


	def export_summary(self):
		ws = self.wb.create_sheet(self.dataset.organization_id.upper() + ' ' + utils.get_pretty_ust_or_release(self.dataset.ust_or_release) + ' Summary')
		conn = utils.connect_db()
		cur = conn.cursor() 
		sql = f"""select summary_item, summary_detail 
				from public.v_{self.dataset.ust_or_release}_control_summary
				where {self.dataset.ust_or_release}_control_id = %s
				order by sort_order"""
		utils.process_sql(conn, cur, sql, params=(self.dataset.control_id,))
		data = cur.fetchall()
		cur.close()
		conn.close()
		for rowno, row in enumerate(data, start=1):
			for colno, cell_value in enumerate(row, start=1):
				ws.cell(row=rowno, column=colno).value = cell_value
				if colno == 2 and ws.cell(row=rowno, column=colno-1).value == 'organization_compartment_flag' and ws.cell(row=rowno, column=colno).value == None:
					 ws.cell(row=rowno, column=colno).fill = utils.get_fill_gen(yellow_cell_fill)
		utils.autowidth(ws)
		ws.column_dimensions['B'].width = 50
		logger.info('Exported %s_control_table summary for control_id %s to %s', self.dataset.ust_or_release, self.dataset.control_id, self.dataset.export_file_path)


	def export_row_counts(self):
		ws = self.wb.create_sheet(self.dataset.organization_id.upper() + ' ' + utils.get_pretty_ust_or_release(self.dataset.ust_or_release) + ' Row Count')
		headers = ['Table Name', 'Number of Rows']
		for colno, header in enumerate(headers, start=1):
			ws.cell(row=1, column=colno).value = header
		header_range = ws["A1:Z1"]
		for row in header_range:
			for cell in row:
				cell.font = Font(bold=True)

		conn = utils.connect_db()
		cur = conn.cursor() 
		sql = f"""select table_name, num_rows 
				from public.v_{self.dataset.ust_or_release}_row_count_summary
				where {self.dataset.ust_or_release}_control_id = %s
				order by sort_order"""
		utils.process_sql(conn, cur, sql, params=(self.dataset.control_id,))
		data = cur.fetchall()
		cur.close()
		conn.close()
		for rowno, row in enumerate(data, start=2):
			for colno, cell_value in enumerate(row, start=1):
				ws.cell(row=rowno, column=colno).value = cell_value
				if colno == 2:
					ws.cell(row=rowno, column=colno).number_format = '#,##0'
		utils.autowidth(ws)
		logger.info('Exported table row counts for control_id %s to %s', self.dataset.control_id, self.dataset.export_file_path)


	def performance_measure_comparison(self):
		ws = self.wb.create_sheet('Performance Measure Comparison')
		conn = utils.connect_db()
		cur = conn.cursor() 
		
		if self.dataset.ust_or_release == 'ust':
			headers = ['Performance Measure', 'Total UST']
			for colno, header in enumerate(headers, start=1):
				ws.cell(row=1, column=colno).value = header
			header_range = ws["A1:Z1"]
			for row in header_range:
				for cell in row:
					cell.font = Font(bold=True)

			sql = """select total_type, total_ust
					from public.v_ust_performance_measures
					where organization_id = %s
					order by sort_order"""
			utils.process_sql(conn, cur, sql, params=(self.dataset.organization_id,))
			data = cur.fetchall()
			for rowno, row in enumerate(data, start=2):
				for colno, cell_value in enumerate(row, start=1):
					ws.cell(row=rowno, column=colno).value = cell_value
					if colno == 2:
						ws.cell(row=rowno, column=colno).number_format = '#,##0'
			ws.cell(row=rowno, column=1).font = Font(bold=True)			
			ws.cell(row=rowno, column=2).font = Font(bold=True)		

			rowno = rowno + 2			
			ws.cell(row=rowno, column=1).value = 'EPA Template'		
			ws.cell(row=rowno, column=1).font = Font(bold=True, underline='single')

			rowno = rowno + 1
			ws.cell(row=rowno, column=1).value = 'EPA Compartment Status'		
			ws.cell(row=rowno, column=1).font = Font(bold=True)
			ws.cell(row=rowno, column=2).value = 'Total UST'		
			ws.cell(row=rowno, column=2).font = Font(bold=True)

			sql = """select "CompartmentStatus", count(*) from public.v_ust_compartment
					where ust_control_id = %s group by "CompartmentStatus" """
			utils.process_sql(conn, cur, sql, params=(self.dataset.control_id,))
			data = cur.fetchall()
			for rowno, row in enumerate(data, start=rowno+1):
				for colno, cell_value in enumerate(row, start=1):
					ws.cell(row=rowno, column=colno).value = cell_value
					if colno == 2:
						ws.cell(row=rowno, column=colno).number_format = '#,##0'

			rowno = rowno + 1
			sql = "select count(*) from public.v_ust_compartment where ust_control_id = %s" 
			utils.process_sql(conn, cur, sql, params=(self.dataset.control_id,))
			cnt = cur.fetchone()[0]
			ws.cell(row=rowno, column=1).value = 'Total UST EPA Template'
			ws.cell(row=rowno, column=1).font = Font(bold=True)
			ws.cell(row=rowno, column=2).value = cnt 
			ws.cell(row=rowno, column=2).font = Font(bold=True)
			ws.cell(row=rowno, column=2).number_format = '#,##0'

		else: # release
			headers = ['Performance Measure', 'Total Releases']
			for colno, header in enumerate(headers, start=1):
				ws.cell(row=1, column=colno).value = header
			header_range = ws["A1:Z1"]
			for row in header_range:
				for cell in row:
					cell.font = Font(bold=True)

			sql = f"""select total_type, total_cumulative_releases
					from public.v_release_performance_measures
					where organization_id = %s
					order by sort_order"""
			utils.process_sql(conn, cur, sql, params=(self.dataset.organization_id,))
			data = cur.fetchall()
			for rowno, row in enumerate(data, start=2):
				for colno, cell_value in enumerate(row, start=1):
					ws.cell(row=rowno, column=colno).value = cell_value
					if colno == 2:
						ws.cell(row=rowno, column=colno).number_format = '#,##0'
			ws.cell(row=rowno, column=1).font = Font(bold=True)			
			ws.cell(row=rowno, column=2).font = Font(bold=True)		

			rowno = rowno + 2
			ws.cell(row=rowno, column=1).value = 'EPA Template'		
			ws.cell(row=rowno, column=1).font = Font(bold=True, underline='single')

			rowno = rowno + 1
			ws.cell(row=rowno, column=1).value = 'Release Status'		
			ws.cell(row=rowno, column=1).font = Font(bold=True)
			ws.cell(row=rowno, column=2).value = 'Total Releases'		
			ws.cell(row=rowno, column=2).font = Font(bold=True)
			sql = f"""select "USTReleaseStatus", count(*) from public.v_ust_release
					where release_control_id = %s group by "USTReleaseStatus" """
			utils.process_sql(conn, cur, sql, params=(self.dataset.control_id,))
			data = cur.fetchall()
			for rowno, row in enumerate(data, start=rowno+1):
				for colno, cell_value in enumerate(row, start=1):
					ws.cell(row=rowno, column=colno).value = cell_value
					if colno == 2:
						ws.cell(row=rowno, column=colno).number_format = '#,##0'	

			rowno = rowno + 1			
			sql = "select count(*) from public.v_ust_release where release_control_id = %s" 
			utils.process_sql(conn, cur, sql, params=(self.dataset.control_id,))
			cnt = cur.fetchone()[0]
			ws.cell(row=rowno, column=1).value = 'Total Releases EPA Template'
			ws.cell(row=rowno, column=1).font = Font(bold=True)
			ws.cell(row=rowno, column=2).value = cnt 
			ws.cell(row=rowno, column=2).font = Font(bold=True)
			ws.cell(row=rowno, column=2).number_format = '#,##0'

		cur.close()
		conn.close()
		
		utils.autowidth(ws)
		logger.info('Exported performance measure summary information for control_id %s to %s', self.dataset.control_id, self.dataset.export_file_path)


	def mapped_element_summary(self):
		ws = self.wb.create_sheet('Mapped Element Summary')
		conn = utils.connect_db()
		cur = conn.cursor() 

		rowno = 1 
		ws.cell(row=rowno, column=1).value = 'Mapped Data Elements'
		ws.cell(row=rowno, column=1).font = Font(bold=True, underline='single')
		ws.cell(row=rowno, column=4).value = 'Unmapped Data Elements'
		ws.cell(row=rowno, column=4).font = Font(bold=True, underline='single')
		
		rowno = 2
		ws.cell(row=rowno, column=1).value = 'Table Name'
		ws.cell(row=rowno, column=1).font = Font(bold=True)
		ws.cell(row=rowno, column=2).value = 'Column Name'
		ws.cell(row=rowno, column=2).font = Font(bold=True)

		ws.cell(row=rowno, column=4).value = 'Table Name'
		ws.cell(row=rowno, column=4).font = Font(bold=True)
		ws.cell(row=rowno, column=5).value = 'Column Name'
		ws.cell(row=rowno, column=5).font = Font(bold=True)

		start = 3
		sql = f"""select table_name, element_name 
				from public.v_{self.dataset.ust_or_release}_element_mapping_for_export
				where {self.dataset.ust_or_release}_control_id = %s
				order by table_sort_order, column_sort_order"""
		utils.process_sql(conn, cur, sql, params=(self.dataset.control_id,))
		data = cur.fetchall()
		for rowno, row in enumerate(data, start=start):
			for colno, cell_value in enumerate(row, start=1):
				ws.cell(row=rowno, column=colno).value = cell_value

		start = 3
		sql = f"""select template_tab_name, element_name from
					(select template_tab_name, element_name, d.sort_order as table_sort_order, c.sort_order as column_sort_order
					from public.{self.dataset.ust_or_release}_elements a join public.{self.dataset.ust_or_release}_elements_tables b on a.element_id = b.element_id 
						join {self.dataset.ust_or_release}_elements_tables c on b.element_id = c.element_id and c.table_name = b.table_name
						join {self.dataset.ust_or_release}_template_data_tables d on b.table_name = d.table_name
					where not exists 
						(select 1 from public.v_{self.dataset.ust_or_release}_element_mapping_for_export x 
						where x.{self.dataset.ust_or_release}_control_id = %s
						and b.table_name = x.epa_table_name and a.database_column_name = x.epa_column_name)) a
				where not (template_tab_name in ('Facility Dispenser','Tank Dispenser','Compartment Dispenser','Piping','Compartment Substance')
							and element_name in ('FacilityID','TankID','TankName','CompartmentID','CompartmentName'))
				and not (template_tab_name in ('Compartment','Tank Substance') and element_name in  ('FacilityID','TankID','TankName'))
				and not (template_tab_name = 'Tank' and element_name = 'FacilityID')
				and not (template_tab_name in ('Substance','Source','Cause','Corrective Action Strategy') and element_name = 'ReleaseID')
				and element_name not like '%%Comment'
				order by table_sort_order, column_sort_order"""
		utils.process_sql(conn, cur, sql, params=(self.dataset.control_id,))
		data = cur.fetchall()
		for rowno, row in enumerate(data, start=start):
			for colno, cell_value in enumerate(row, start=4):
				ws.cell(row=rowno, column=colno).value = cell_value

		cur.close()
		conn.close()				
		utils.autowidth(ws)
		logger.info('Exported mapped element summary for control_id %s to %s', self.dataset.control_id, self.dataset.export_file_path)



def main(ust_or_release, organization_id=None, control_id=None, export_file_name=None, export_file_dir=None, export_file_path=None):
	dataset = Dataset(ust_or_release=ust_or_release,
		              control_id=control_id,
		              base_file_name='control_summary_' + utils.get_timestamp_str() + '.xlsx',
					  export_file_name=export_file_name,
					  export_file_dir=export_file_dir,
					  export_file_path=export_file_path)

	template = Summary(dataset=dataset)


if __name__ == '__main__':   
	main(ust_or_release=ust_or_release,
		 control_id=control_id, 
		 export_file_name=export_file_name,
		 export_file_dir=export_file_dir,
		 export_file_path=export_file_path)
