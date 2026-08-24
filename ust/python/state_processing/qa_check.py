
import openpyxl as op
import psycopg2.errors
from openpyxl.styles import Font

from ust.python.state_processing import element_mapping_to_excel
from ust.python.state_processing.qa_exclusions import Exclusions
from ust.python.state_processing.qa_summary_counts import SummaryCounts
from ust.python.util import utils
from ust.python.util.dataset import Dataset
from ust.python.util.logger_factory import logger

ust_or_release = ''             # Valid values are 'ust' or 'release'
control_id = 0                    # Enter an integer that is the ust_control_id or release_control_id
organization_id = ''            # Optional; only used if control_id is not passed. If control_id == 0 or None, the script will retrieve the most recent control_id for the organization. 
force_exclusions = False        # Boolean; defaults to False. If False, will only generate exclusions (e.g. unregulated substances, etc.) if there are no errors. Set to True to force exclusion export even if there are errors to resolve.
force_summary_counts = False    # Boolean; defaults to False. If False, will only generate summary counts if there are no errors. Set to True to force summary counts even if there are errors to resolve.
include_details = True          # Boolean; defaults to True. Set to False to skip detail worksheets and speed up QA runs.

# These variables can usually be left unset. This script will generate an Excel spreadsheet in the appropriate state folder in the repo under /ust/python/exports/QAQC
# This file directory and its contents are excluded from pushes to the repo by .gitignore.
export_file_path = None
export_file_dir = None
export_file_name = None

join_cols = {}
join_cols['v_ust_facility'] = []
join_cols['v_ust_facility_dispenser'] = ['facility_id']
join_cols['v_ust_tank'] = ['facility_id']
join_cols['v_ust_tank_substance'] = ['facility_id','tank_id']
join_cols['v_ust_tank_dispenser'] = ['facility_id','tank_id']
join_cols['v_ust_compartment'] = ['facility_id','tank_id']
join_cols['v_ust_compartment_substance'] = ['facility_id','tank_id','compartment_id']
join_cols['v_ust_compartment_dispenser'] = ['facility_id','tank_id','compartment_id']
join_cols['v_ust_piping'] = ['facility_id','tank_id','compartment_id']
join_cols['v_ust_release'] = []
join_cols['v_ust_release_source'] = ['release_id'] 
join_cols['v_ust_release_cause'] = ['release_id']
join_cols['v_ust_release_substance'] = ['release_id']
join_cols['v_ust_release_corrective_action_strategy'] = ['release_id']

yellow_cell_fill = 'FFFF00' # yellow


class QualityCheck:
    conn = None 
    cur = None
    wb = None
    view_name = None 
    table_name = None 
    view_col_str = None 

    def __init__(self, dataset, force_exclusions=False, force_summary_counts=False, include_details=True):
        self.dataset = dataset
        self.force_exclusions = force_exclusions
        self.force_summary_counts = force_summary_counts
        self.include_details = include_details
        self.views_to_review = []
        self.error_dict = {}
        self.error_cnt_dict = {}
        self.view_counts = {}
        self.view_columns_cache = {}
        self.header_cache = {}
        self.required_nonnull_cols_cache = {}
        self.key_cols_cache = {}
        self.check_constraints_cache = {}
        self.lookup_values_cache = {}
        self.relation_columns_cache = {}


    def _get_view_columns(self, view_name):
        if view_name in self.view_columns_cache:
            return self.view_columns_cache[view_name]
        sql = """select column_name
                from information_schema.columns
                where table_schema = %s and table_name = %s
                order by ordinal_position"""
        utils.process_sql(self.conn, self.cur, sql, params=(self.dataset.schema, view_name))
        columns = [r[0] for r in self.cur.fetchall()]
        self.view_columns_cache[view_name] = columns
        return columns


    def _get_relation_columns(self, table_name, schema=None):
        schema = schema or self.dataset.schema
        cache_key = (schema, table_name)
        if cache_key in self.relation_columns_cache:
            return self.relation_columns_cache[cache_key]
        sql = """select column_name
                from information_schema.columns
                where table_schema = %s and table_name = %s
                order by ordinal_position"""
        utils.process_sql(self.conn, self.cur, sql, params=(schema, table_name))
        columns = [r[0] for r in self.cur.fetchall()]
        self.relation_columns_cache[cache_key] = columns
        return columns


    def _relation_has_column(self, table_name, column_name, schema=None):
        return column_name in self._get_relation_columns(table_name, schema=schema)


    def _quote_ident(self, name):
        return '"' + name.replace('"', '""') + '"'


    def _get_required_nonnull_cols(self, table_name):
        if table_name in self.required_nonnull_cols_cache:
            return self.required_nonnull_cols_cache[table_name]
        sql = """select column_name from information_schema.columns 
                where table_schema = 'public' and table_name = %s 
                and is_nullable = 'NO' and ordinal_position > 1
                and column_name not like 'ust%%id' and column_name not like 'release%%id'
                order by ordinal_position"""
        utils.process_sql(self.conn, self.cur, sql, params=(table_name,))
        cols = [row[0] for row in self.cur.fetchall()]
        self.required_nonnull_cols_cache[table_name] = cols
        return cols


    def _get_key_cols(self, view_name):
        if view_name in self.key_cols_cache:
            return self.key_cols_cache[view_name]
        sql = f"""select column_name from public.{self.dataset.ust_or_release}_view_key_columns 
                  where view_name = %s order by sort_order"""
        utils.process_sql(self.conn, self.cur, sql, params=(view_name,))
        cols = [r[0] for r in self.cur.fetchall()]
        self.key_cols_cache[view_name] = cols
        return cols


    def _get_check_constraints(self, table_name):
        if table_name in self.check_constraints_cache:
            return self.check_constraints_cache[table_name]
        sql = """select cc.constraint_name, check_clause
                    from information_schema.check_constraints cc 
                        join pg_constraint cons on cc.constraint_name = cons.conname
                        join pg_class t on cons.conrelid = t.oid 
                    where constraint_schema = 'public' and t.relname = %s
                    order by 1, 2"""
        utils.process_sql(self.conn, self.cur, sql, params=(table_name,))
        rows = self.cur.fetchall()
        self.check_constraints_cache[table_name] = rows
        return rows



    def process(self):
        self.connect_db()
        self.set_views()
        if not self.views_to_review:
            logger.warning('No %s template views found in schema %s; exiting.', self.dataset.ust_or_release, self.dataset.schema)
            logger.info('Views this script looks for: %s', self.get_view_names())
            self.disconnect_db()
            raise RuntimeError(f'No {self.dataset.ust_or_release} template views found in schema {self.dataset.schema}.')
        self.wb = op.Workbook()    
        self.check_missing_views()
        self.set_view_counts()
        self.check_view_counts()
        for view_name in self.views_to_review:
            self.view_name = view_name
            self.table_name = view_name.replace('v_','')
            self.set_view_col_str()
            self.check_join_cols()
            self.check_required_cols()
            self.check_duplicate_rows()
            self.check_extraneous_cols()
            self.check_nonunique()
            self.check_bad_datatypes()
            self.check_failed_constraints()
            self.check_missing_mapping()
            self.check_wrong_mapping_cols()
            self.check_bad_mapping()
            if self.dataset.ust_or_release == 'ust':
                self.check_compartment_data_flag()
            self.check_unregulated_parents()
            self.check_missing_parent_view_keys()
        # self.check_inactive_substances()    # this is now covered under check_substance_types
        self.check_substance_types()
        self.check_unregulated_substances()
        self.write_overview()
        self.exclusions()
        self.summary_counts()
        element_mapping_to_excel.build_ws(self.dataset, self.wb.create_sheet(), admin=True)
        self.cleanup_wb()
        self.disconnect_db()
        if not self.error_dict:
            logger.info('\nAll QAQC checks passed!!\n')     


    def connect_db(self):
        if not self.conn:
            self.conn = utils.connect_db()
            self.cur = self.conn.cursor()
            logger.info('Connected to database')
        

    def disconnect_db(self):
        if self.conn:
            self.cur.close()
            self.conn.close()
            self.conn = None 
            logger.info('Disconnected from database')


    def get_view_names(self):
        sql = f"select view_name from public.{self.dataset.ust_or_release}_template_data_tables order by sort_order"
        utils.process_sql(self.conn, self.cur, sql)
        rows = self.cur.fetchall()
        views = [r[0] for r in rows]
        return views 


    def set_views(self):
        sql = f"""select a.table_name as view_name 
                    from information_schema.tables a join public.{self.dataset.ust_or_release}_template_data_tables b on a.table_name = b.view_name 
                    where a.table_schema = %s
                    order by b.sort_order"""
        utils.process_sql(self.conn, self.cur, sql, params=(self.dataset.schema,))
        rows = self.cur.fetchall()        
        self.views_to_review = [r[0] for r in rows]
        logger.info("The following views will be QC'ed: %s", self.views_to_review)


    def set_view_counts(self):
        for view_name in self.views_to_review:
            sql = f"select count(*) from {self.dataset.schema}.{view_name}"
            try:
                self.cur.execute(sql)
            except psycopg2.errors.UndefinedTable:
                continue
            num_rows = self.cur.fetchone()[0]
            self.view_counts[view_name] = num_rows


    def check_view_counts(self):
        if (
            self.dataset.ust_or_release == 'ust'
            and 'v_ust_compartment' in self.views_to_review
            and 'v_ust_tank' in self.views_to_review
            and self.view_counts['v_ust_compartment'] < self.view_counts['v_ust_tank']
        ):
            self.error_dict['Fewer rows in child table than expected'] = 'v_compartment_tank should have at least as many rows as v_ust_tank'
            # if 'v_ust_piping' in self.views_to_review and 'v_ust_compartment' in self.views_to_review:
            #     if self.view_counts['v_ust_piping'] < self.view_counts['v_ust_compartment']:
            #         self.error_dict['Fewer rows in child table than expected'] = 'v_ust_piping should have at least as many rows as v_ust_compartment'


    def check_missing_views(self):
        # check that all parent 
        missing_views = []
        if self.dataset.ust_or_release == 'ust':
            if 'v_ust_piping' in self.views_to_review:
                if 'v_ust_compartment' not in self.views_to_review:
                    missing_views.insert(0, 'v_ust_compartment')
                if 'v_ust_tank' not in self.views_to_review:
                    missing_views.insert(0, 'v_ust_tank')
            if 'v_ust_compartment' in self.views_to_review and 'v_ust_tank' not in self.views_to_review and 'v_ust_tank' not in missing_views:
                missing_views.insert(0, 'v_ust_tank')
            for view_name in missing_views:
                self.error_dict['Missing required view (child data present)'] = self.dataset.schema + '.' + view_name  
                logger.warning('Missing required view (child data present) %s.%s', self.dataset.schema, view_name)


    def set_view_col_str(self):
        rows = [(c,) for c in self._get_view_columns(self.view_name)]
        col_str = ''
        for row in rows:
            col_str = col_str + row[0] + ', '
        if col_str:
            col_str = col_str[:-2]
        self.view_col_str = col_str 


    def _select_count(self, sql, params=None):
        utils.process_sql(self.conn, self.cur, sql, params=params)
        return self.cur.fetchone()[0]


    def write_to_ws(self, data, ws_name):
        if not self.include_details:
            return
        ws_name = ws_name[:31]
        if data:
            ws = self.wb.create_sheet(ws_name)
            header_key = (self.dataset.schema, self.view_name)
            headers = self.header_cache.get(header_key)
            if headers is None:
                headers = utils.get_headers(self.view_name, self.dataset.schema)
                self.header_cache[header_key] = headers
            ws.append(headers)
            for row in data:
                ws.append(list(row))
            logger.info('Data written to worksheet %s', ws_name)
        else:
            logger.info('Nothing to write to %s', ws_name)


    def write_invalid_epa_values_to_ws(self, data):
        if not self.include_details:
            return
        ws_name = 'Invalid EPA values'
        headers = ['EPA Table', 'EPA Column', 'Invalid EPA Value', 'Lookup Table', 'Lookup Column', 'Valid EPA Values']
        if ws_name in self.wb.sheetnames:
            ws = self.wb[ws_name]
        else:
            ws = self.wb.create_sheet(ws_name)
            ws.append(headers)
        for row in data:
            ws.append(list(row))
        utils.autowidth(ws)
        logger.info('Data written to worksheet %s', ws_name)


    def check_join_cols(self):
        # check for missing columns in the view that join child to parent tables 
        req_cols = join_cols[self.view_name]
        existing_cols = self._get_view_columns(self.view_name)
        for rcol in req_cols:
            if rcol not in existing_cols:
                self.error_dict['Missing join column'] = self.dataset.schema + '.' + self.view_name + '.' + rcol 
                logger.warning('Missing join column %s in view %s.%s', rcol, self.dataset.schema, self.view_name)


    def check_required_cols(self):
        # check for missing columns in the view that are required by EPA 
        existing_cols = set(self._get_view_columns(self.view_name))

        rows = [(col_name,) for col_name in self._get_required_nonnull_cols(self.table_name)]
        present_required_cols = []
        for row in rows:
            col_name = row[0]
            if col_name not in existing_cols:
                self.error_dict['Missing required column'] = self.dataset.schema + '.' + self.view_name + '.' + col_name 
                logger.warning('Missing required column %s in view %s.%s', col_name, self.dataset.schema, self.view_name)
            else:
                present_required_cols.append(col_name)

        if not present_required_cols:
            return

        aggregate_terms = [
            f"sum(case when {self._quote_ident(col_name)} is null then 1 else 0 end)"
            for col_name in present_required_cols
        ]
        aggregate_sql = (
            f"select {', '.join(aggregate_terms)} "
            f"from {self._quote_ident(self.dataset.schema)}.{self._quote_ident(self.view_name)}"
        )
        utils.process_sql(self.conn, self.cur, aggregate_sql)
        null_counts = self.cur.fetchone()

        for col_name, num_rows in zip(present_required_cols, null_counts):
            num_rows = int(num_rows or 0)
            self.error_cnt_dict['Number of null rows for required column ' + self.table_name + '.' + col_name] = num_rows
            logger.warning('Number of null rows for required column %s.%s = %s', self.table_name, col_name, num_rows)
            if num_rows > 0 and self.include_details:
                details_sql = f"select * from {self.dataset.schema}.{self.view_name} where {self._quote_ident(col_name)} is null"
                utils.process_sql(self.conn, self.cur, details_sql)
                self.write_to_ws(self.cur.fetchall(), col_name + ' null')


    def check_duplicate_rows(self):
        # check for rows that have duplicate key columns
        key_cols = self._get_key_cols(self.view_name)
        key_col_str = ''
        join = ''
        for col in key_cols:
            key_col_str = key_col_str + col + ', '
            join = join + 'a.' + col + ' = b.' + col + ' and ' 
        key_col_str = key_col_str[:-2]
        join = join[:-4]
        sql = f"""select {key_col_str}, count(*) num_rows from {self.dataset.schema}.{self.view_name} 
                  group by {key_col_str} having count(*) > 1"""
        utils.process_sql(self.conn, self.cur, sql)
        rows = self.cur.fetchall()
        num_rows = len(rows) 
        self.error_cnt_dict['Number of duplicated key columns in ' + self.dataset.schema + '.' + self.view_name + ' (' + key_col_str + ')'] = num_rows
        logger.warning('Number of duplicated key columns in %s.%s: %s', self.dataset.schema, self.view_name, num_rows)
        if num_rows > 0:
            if self.include_details:
                sql = f"""select * from {self.dataset.schema}.{self.view_name}  a
                        where exists
                            (select {key_col_str}
                            from {self.dataset.schema}.{self.view_name}  b
                            where {join}
                            group by {key_col_str}
                            having count(*) > 1)
                        order by 1, 2, 3"""
                utils.process_sql(self.conn, self.cur, sql)
                data = self.cur.fetchall()
                self.write_to_ws(data, self.view_name + ' duplicates')
            num_rows = len(rows) 
            self.error_dict['Number of rows with duplicated key columns in ' + self.dataset.schema + '.' + self.view_name ] = num_rows


    def get_bad_datatypes(self, data):
        for d in data:
            col_name = d[0]
            table_data_type = d[1]
            table_len = d[2]
            view_data_type = d[3]
            view_len = d[4]
            if view_len and table_len and view_len > table_len:
                count_sql = f"select count(*) from {self.dataset.schema}.{self.view_name} where length({col_name}) > %s"
                num_rows = self._select_count(count_sql, params=(table_len,))
                self.error_cnt_dict['Number of rows exceeding allowed length of ' + self.table_name + '.' + col_name] = num_rows
                logger.warning('Number of rows exceeding allowed length of %s.%s: %s', self.table_name, col_name, num_rows)
                if num_rows > 0 and self.include_details:
                    sql2 = f"select * from {self.dataset.schema}.{self.view_name} where length({col_name}) > %s"
                    utils.process_sql(self.conn, self.cur, sql2, params=(table_len,))
                    self.write_to_ws(self.cur.fetchall(), col_name + ' too long')
            elif view_data_type == 'text' and table_data_type == 'character varying':
                count_sql = f"select count(*) from {self.dataset.schema}.{self.view_name} where length({col_name}) > %s"
                num_rows = self._select_count(count_sql, params=(table_len,))
                if num_rows > 0:
                    self.error_cnt_dict['Number of rows exceeding allowed length of ' + self.table_name + '.' + col_name] = num_rows
                    logger.warning('Number of rows exceeding allowed length of %s.%s: %s', self.table_name, col_name, num_rows)
                    if self.include_details:
                        sql = f"select * from {self.dataset.schema}.{self.view_name} where length({col_name}) > %s"
                        utils.process_sql(self.conn, self.cur, sql, params=(table_len,))
                        self.write_to_ws(self.cur.fetchall(), col_name + ' too long')
            elif table_data_type != view_data_type:
                self.error_dict['Wrong data type for ' + self.table_name + '.' + col_name] = self.dataset.schema + '.' + self.view_name + '.' + col_name


    def check_bad_datatypes(self):
        # check for columns in the state schema view where the data type doesn't match the EPA table, or 
        # the length of the value in a character column is too long in the state data 
        sql = """select a.column_name, a.data_type, a.character_maximum_length, b.data_type, b.character_maximum_length 
                from information_schema.columns a join information_schema.columns b on a.column_name = b.column_name
                where a.table_schema = 'public' and a.table_name = %s
                and b.table_schema  = %s and b.table_name = %s
                and (a.data_type <> b.data_type or b.character_maximum_length > a.character_maximum_length)
                order by a.ordinal_position"""
        utils.process_sql(self.conn, self.cur, sql, params=(self.table_name, self.dataset.schema, self.view_name))
        data = self.cur.fetchall()
        self.get_bad_datatypes(data)

        # check the data type of the EPA table join columns 
        if self.dataset.ust_or_release == 'ust':
            if self.view_name == 'v_ust_tank':
                epa_table_name = 'ust_facility'
            elif self.view_name == 'v_ust_compartment':
                epa_table_name = 'ust_tank'
            elif self.view_name == 'v_ust_piping':
                epa_table_name = 'ust_compartment'
            else: # no need to check v_ust_facility as it is the parent
                epa_table_name = None 
            if epa_table_name:
                sql = """select a.column_name, a.data_type, a.character_maximum_length, b.data_type, b.character_maximum_length 
                        from information_schema.columns a join information_schema.columns b on a.column_name = b.column_name
                        where a.table_schema = 'public' and a.table_name = %s
                        and b.table_schema = %s and b.table_name = %s
                        and (a.data_type <> b.data_type or b.character_maximum_length > a.character_maximum_length)
                        order by a.ordinal_position"""
                utils.process_sql(self.conn, self.cur, sql, params=(epa_table_name, self.dataset.schema, self.view_name))
                data = self.cur.fetchall()
                self.get_bad_datatypes(data)


    def check_extraneous_cols(self):
        # check for columns in the state schema view that don't correspond to columns in the EPA template 
        sql = """select column_name from information_schema.columns 
                where table_schema = %s and table_name = %s and column_name not in 
                    (select column_name from information_schema.columns 
                    where table_schema = 'public' and table_name = %s)
                order by column_name"""
        utils.process_sql(self.conn, self.cur, sql, params=(self.dataset.schema, self.view_name, self.table_name))
        rows = self.cur.fetchall()
        for row in rows:
            col_name = row[0]
            if col_name not in join_cols[self.view_name] and not (self.view_name == 'v_ust_compartment_substance' and col_name == 'substance_id'):
                self.error_dict['Extraneous column'] = self.dataset.schema + '.' + self.view_name + '.' + col_name 
                logger.warning('Extraneous column %s in view %s.%s', col_name, self.dataset.schema, self.view_name)


    def check_nonunique(self):
        # check for non-unique (repeating) rows    
        if not self.include_details:
            logger.info('Skipping full-row nonunique check for %s.%s in fast QA mode.', self.dataset.schema, self.view_name)
            self.error_cnt_dict['nonunique rows in ' + self.dataset.schema + '.' + self.view_name] = 0
            return
        count_sql = f"""select count(*)
                        from (
                            select 1 from {self.dataset.schema}.{self.view_name}
                            group by {self.view_col_str}
                            having count(*) > 1
                        ) x"""
        num_rows = self._select_count(count_sql)
        self.error_cnt_dict['nonunique rows in ' + self.dataset.schema + '.' + self.view_name] = num_rows
        logger.warning('Number of non-unique rows in %s.%s: %s', self.dataset.schema, self.view_name, num_rows)
        if num_rows > 0 and self.include_details:
            sql = f"select {self.view_col_str}, count(*) from {self.dataset.schema}.{self.view_name} group by {self.view_col_str} having count(*) > 1 order by 1, 2"
            utils.process_sql(self.conn, self.cur, sql)
            self.write_to_ws(self.cur.fetchall(), self.view_name + ' nonunique')


    def check_failed_constraints(self):
        # check for failed check constraints
        rows = self._get_check_constraints(self.table_name)
        for row in rows:
            constraint_name = row[0]
            check_clause = row[1]
            sql2 = f"select count(*) from {self.dataset.schema}.{self.view_name} where not {check_clause}"
            try:
                self.cur.execute(sql2)
            except psycopg2.errors.UndefinedColumn:
                continue 
            except psycopg2.Error as e:
                logger.error('Error processing SQL: %s', e)
                utils.pretty_print_query(self.cur)
                self.conn.rollback()
                self.cur.close()
                self.conn.close()        
                raise RuntimeError(f'Failed evaluating check constraint {constraint_name} against {self.dataset.schema}.{self.view_name}.') from e
            num_rows = self.cur.fetchone()[0]
            self.error_cnt_dict['failed check constraint ' + self.dataset.schema + '.' + constraint_name] = num_rows
            logger.warning('Number of failed rows for check constraint %s.%s: %s', self.table_name, constraint_name, num_rows)
            if num_rows > 0 and self.include_details:
                detail_sql = f"select * from {self.dataset.schema}.{self.view_name} where not {check_clause}"
                utils.process_sql(self.conn, self.cur, detail_sql)
                self.write_to_ws(self.cur.fetchall(), constraint_name)


    def check_missing_mapping(self):
        sql = f"""select c.column_name 
                from information_schema.columns c 
                    join information_schema.tables t 
                        on c.table_schema = t.table_schema and c.table_name = t.table_name
                    join public.{self.dataset.ust_or_release}_template_data_tables x on c.table_name = x.view_name
                where c.table_schema = %s and c.table_name = %s 
                and column_name not in ('facility_state', 'facility_epa_region', 'epa_region', 'state') and not exists 
                    (select 1 from public.{self.dataset.ust_or_release}_element_mapping m
                    where x.table_name = m.epa_table_name and c.column_name = m.epa_column_name
                    and m.{self.dataset.ust_or_release}_control_id = %s)
                order by c.ordinal_position"""
        utils.process_sql(self.conn, self.cur, sql, params=(self.dataset.schema, self.view_name, self.dataset.control_id))
        rows = self.cur.fetchall()
        num_rows = len(rows) 
        self.error_cnt_dict['Unmapped elements in ' + self.dataset.schema + '.' + self.view_name] = num_rows
        unmapped_cols = ''
        for row in rows:
            unmapped_cols = row[0] + '; '
        if unmapped_cols:
            unmapped_cols = unmapped_cols[:-2]
            self.error_dict['Unmapped elements in ' + self.view_name] = unmapped_cols 
            logger.warning('Unmapped elements in %s: %s', self.view_name, unmapped_cols)
            

    def check_wrong_mapping_cols(self):
        # check for mapping of description column names instead of ID column names (e.g. "substance" instead of "substance_id")
        sql = f"""select epa_table_name, epa_column_name  
                from public.{self.dataset.ust_or_release}_element_mapping a
                where {self.dataset.ust_or_release}_control_id = %s
                and epa_column_name not in ('tank_name','compartment_name') 
                and not exists 
                    (select 1 from public.v_{self.dataset.ust_or_release}_element_metadata b
                    where a.epa_table_name = b.table_name and a.epa_column_name = b.column_name)
                order by 1, 2"""
        utils.process_sql(self.conn, self.cur, sql, params=(self.dataset.control_id,))
        rows = self.cur.fetchall()
        num_rows = len(rows) 
        self.error_cnt_dict['Wrong elements mapped in ' + self.dataset.ust_or_release + '_element_mapping'] = num_rows
        for row in rows:
            table = row[0]
            col = row[1]
            self.error_dict['Wrong element mapped in table ' + table + '; perhaps you meant to map ' + col + '_id?'] = col 
            logger.warning('Wrong element %s mapped for table %s; perhaps you meant to map %s_id?', col, table, col)
            

    # def check_inactive_substances(self):
    #     # Check that any substances mapped are flagged as inactive in the lookup table 
    #     sql = f"""select distinct epa_table_name, epa_value 
    #             from public.v_{self.dataset.ust_or_release}_element_mapping a join public.substances b on a.epa_value = b.substance 
    #             where {self.dataset.ust_or_release}_control_id = %s 
    #             and a.epa_column_name = 'substance_id' and b.inactive_flag is not null
    #             order by 1, 2"""
    #     utils.process_sql(self.conn, self.cur, sql, params=(self.dataset.control_id,))
    #     rows = self.cur.fetchall()
    #     num_rows = len(rows) 
    #     self.error_cnt_dict['Inactive EPA substances mapped in ' + self.dataset.ust_or_release + '_element_value_mapping'] = num_rows
    #     for row in rows:
    #         epa_table_name = row[0]
    #         epa_value = row[1]
    #         self.error_dict['Inactive EPA substance mapped in ' + epa_table_name] = epa_value
    #         logger.warning('Inactive EPA substance "%s" mapped in %s_element_value_mapping', epa_value, epa_table_name)


    def check_substance_types(self):
        # Check that any substances are flagged as UST/Release as appropriate in Substances lookup table
        sql = f"""select distinct epa_table_name, epa_value 
                from public.v_{self.dataset.ust_or_release}_element_mapping a join public.substances b on a.epa_value = b.substance 
                where {self.dataset.ust_or_release}_control_id = %s 
                and a.epa_column_name = 'substance_id' and b.{self.dataset.ust_or_release}_flag is null
                order by 1, 2"""
        utils.process_sql(self.conn, self.cur, sql, params=(self.dataset.control_id,))
        rows = self.cur.fetchall()
        num_rows = len(rows) 
        subtype = 'Prevention'
        if self.dataset.ust_or_release == 'release':
            subtype = 'Cleanup'
        self.error_cnt_dict['Substance not valid for ' + subtype + ' mapped in ' + self.dataset.ust_or_release + '_element_value_mapping'] = num_rows
        for row in rows:
            epa_table_name = row[0]
            epa_value = row[1]
            self.error_dict['Substance not valid for ' + subtype + ' mapped in ' + epa_table_name] = epa_value
            logger.warning('Invalid %s substance "%s" mapped in %s_element_value_mapping', subtype, epa_value, epa_table_name)


    def check_bad_mapping(self):
        # check for bad mapping values
        sql = f"""select distinct epa_column_name, epa_value, database_lookup_table, database_column_name 
                from public.v_{self.dataset.ust_or_release}_element_mapping a join public.{self.dataset.ust_or_release}_elements b on a.epa_column_name = b.database_column_name 
                where {self.dataset.ust_or_release}_control_id = %s and database_lookup_table is not null and epa_table_name = %s and epa_value is not null
                order by 1, 2, 3"""
        utils.process_sql(self.conn, self.cur, sql, params=(self.dataset.control_id, self.table_name))
        rows = self.cur.fetchall()
        num_errors = 0
        invalid_rows = []
        error_count_key = 'Invalid EPA values in ' + self.dataset.ust_or_release + '_element_value_mapping'
        for row in rows:
            epa_column_name = row[0]
            epa_value = row[1]
            lookup_table = row[2]
            lookup_column = row[3].replace('_id','')
            if lookup_column == 'facility_type1' or lookup_column == 'facility_type2':
                lookup_column = 'facility_type'
            lookup_key = (lookup_table, lookup_column)
            valid_values = self.lookup_values_cache.get(lookup_key)
            if valid_values is None:
                sql2 = (
                    f"select {self._quote_ident(lookup_column)} "
                    f"from public.{self._quote_ident(lookup_table)}"
                )
                utils.process_sql(self.conn, self.cur, sql2)
                valid_values = {value_row[0] for value_row in self.cur.fetchall()}
                self.lookup_values_cache[lookup_key] = valid_values

            if epa_value not in valid_values:
                self.error_dict[f'Invalid EPA value in {epa_column_name}: {epa_value}'] = f'{lookup_table}.{lookup_column}'
                logger.warning('Invalid EPA value for %s.%s: %s', self.table_name, epa_column_name, epa_value)
                num_errors += 1
                invalid_rows.append((
                    self.table_name,
                    epa_column_name,
                    epa_value,
                    lookup_table,
                    lookup_column,
                    ', '.join(str(value) for value in sorted(valid_values)),
                ))
        if num_errors > 0 and self.include_details:
            self.write_invalid_epa_values_to_ws(invalid_rows)
        self.error_cnt_dict[error_count_key] = self.error_cnt_dict.get(error_count_key, 0) + num_errors


    def check_unregulated_substances(self):
        # check for inclusion of tanks/releases that are unregulated due to heating oil
        if self.dataset.ust_or_release == 'ust':
            unregulated_table = 'erg_unregulated_tanks'
            row_type = 'tanks'
            extramsg = ' or small tank at farm/residence'
            sql = """select count(*) from information_schema.columns 
                      where table_schema = %s and table_name = %s and column_name like 'facility_type%%'"""
            utils.process_sql(self.conn, self.cur, sql, params=(self.dataset.schema, 'v_ust_facility'))
            cnt = self.cur.fetchone()[0]
            if cnt == 0:
                logger.info('No facility type column in %s.v_ust_facility so not performating heating oil tank check', self.dataset.schema)
                return 

            sql = f"""select distinct facility_id, tank_id 
                    from 
                        (select ts.facility_id, tank_id 
                        from {self.dataset.schema}.v_ust_tank_substance ts join public.substances s on ts.substance_id = s.substance_id 
                            join (select distinct facility_id from 
                                    (select facility_id, facility_type1 as facility_type_id from {self.dataset.schema}.v_ust_facility ) x 
                                  where facility_type_id <> 4) f on ts.facility_id = f.facility_id
                        where s.substance like 'Heating%'"""
            sql2 = """select count(*) from information_schema.columns 
                      where table_schema = %s and table_name = 'v_ust_compartment' 
                      and column_name = 'compartment_capacity_gallons' """
            utils.process_sql(self.conn, self.cur, sql2, params=(self.dataset.schema,))
            cnt = self.cur.fetchone()[0]
            if cnt > 0:
                sql = sql + f"""\nunion all 
                        select x.facility_id, x.tank_id 
                        from (select facility_id, tank_id, sum(compartment_capacity_gallons) as tank_capacity_gallons 
                              from {self.dataset.schema}.v_ust_compartment group by facility_id, tank_id) x 
                            join (select distinct facility_id from 
                                    (select facility_id, facility_type1 as facility_type_id from {self.dataset.schema}.v_ust_facility ) x 
                                  where facility_type_id in (1,12)) f on x.facility_id = f.facility_id      --Agricultural/farm; Residential
                            join {self.dataset.schema}.v_ust_tank_substance ts on x.facility_id = ts.facility_id and x.tank_id = ts.tank_id
                            join public.substances s on ts.substance_id = s.substance_id
                        where tank_capacity_gallons < 1100 and s.substance_group in ('Diesel','Gasoline') """
            sql = sql + """) a
                    order by 1, 2"""
        else: # releases
            unregulated_table = 'erg_unregulated_substances'
            row_type = 'substances'
            extramsg = ''
            sql = """select count(*) from information_schema.columns 
                      where table_schema = %s and table_name = %s and column_name like 'facility_type%%'"""
            utils.process_sql(self.conn, self.cur, sql, params=(self.dataset.schema, 'v_ust_release'))
            cnt = self.cur.fetchone()[0]
            if cnt == 0:
                logger.info('No facility type column in %s.v_ust_release so not performating heating oil tank check', self.dataset.schema)
                return 

            sql = """select count(*) from information_schema.columns 
                      where table_schema = %s and table_name = %s and column_name = %s"""
            utils.process_sql(self.conn, self.cur, sql, params=(self.dataset.schema, 'v_ust_release_substance', 'substance_id'))
            cnt = self.cur.fetchone()[0]
            if cnt == 0:
                logger.info('No view %s.v_ust_release_substance so not performating heating oil tank check', self.dataset.schema)
                return 

            sql = f"""select distinct ts.release_id
                        from (select release_id from {self.dataset.schema}.v_ust_release where facility_type_id <> 4) r     --Bulk plant storage/petroleum distributor
                            join {self.dataset.schema}.v_ust_release_substance ts on ts.release_id = r.release_id
                            join public.substances s on ts.substance_id = s.substance_id 
                        where s.substance_group = 'Heating'
                    order by 1"""
        utils.process_sql(self.conn, self.cur, sql)
        rows = self.cur.fetchall()
        num_rows = len(rows) 
        unreg_source_view = 'vw_erg_unreg_substances'
        unreg_source_cnt = None

        if utils.get_table_existence(unreg_source_view, self.dataset.schema):
            source_sql = f"select count(*) from {self.dataset.schema}.{unreg_source_view}"
            utils.process_sql(self.conn, self.cur, source_sql)
            unreg_source_cnt = self.cur.fetchone()[0]
        
        if num_rows > 0:
            msg = f'Number of {row_type} that need to be excluded due to unregulated heating oil{extramsg}. '
            if self._relation_has_column(unregulated_table, 'unregulated_reason'):
                sql = f"""select count(*) from {self.dataset.schema}.{unregulated_table} 
                          where lower(unregulated_reason) in ('heating oil','small tank at farm/residence')"""
                utils.process_sql(self.conn, self.cur, sql)
                unreg_cnt = self.cur.fetchone()[0]
                unreg_count_description = f'{self.dataset.schema}.{unregulated_table} rows by reason={unreg_cnt}'
            else:
                sql = f"select count(*) from {self.dataset.schema}.{unregulated_table}"
                utils.process_sql(self.conn, self.cur, sql)
                unreg_cnt = self.cur.fetchone()[0]
                unreg_count_description = f'{self.dataset.schema}.{unregulated_table} rows={unreg_cnt}; missing unregulated_reason column'
            msg += f'QA candidates={num_rows}; {self.dataset.schema}.{unreg_source_view} rows={unreg_source_cnt}; {unreg_count_description}. '
            if unreg_cnt == len(rows):
                msg += f'It looks like exclude_unregulated.py was run but unregulated {row_type} were not excluded from data views.'
            elif unreg_cnt == 0:
                if unreg_source_cnt == 0:
                    msg += (
                        f'{unregulated_table} does not contain expected rows and helper view {unreg_source_view} is empty. '
                        'Run create-unreg with --views-only, then rerun populate-unreg.'
                    )
                else:
                    msg += f'{unregulated_table} does not contain expected rows. Run script exclude_unregulated.py and rebuild data views.'
            else:
                msg += f'{unreg_cnt} rows found in {unregulated_table}. Investigate this discrepancy and rebuild data views.'
            if not self._relation_has_column(unregulated_table, 'unregulated_reason'):
                msg += f' Recreate {self.dataset.schema}.{unregulated_table} with create-unreg --drop-existing so QA can report exclusion reasons.'

        self.error_cnt_dict['Rows with unregulated substances not excluded from views'] = num_rows
        if num_rows > 0:
            logger.warning(msg)
            self.error_dict[msg] = num_rows


    def check_unregulated_parents(self):
        def view_has_column(column_name):
            return column_name in self._get_view_columns(self.view_name)

        if self.dataset.ust_or_release == 'ust':
            if not view_has_column('facility_id'):
                msg = f'Missing column {self.view_name}.facility_id; skipping unregulated parent check.'
                logger.warning(msg)
                self.error_cnt_dict['Rows with unregulated rows not excluded from ' + self.view_name] = 0
                self.error_dict[msg] = 'Skipped'
                return

            if self.view_name == 'v_ust_facility':
                unreg_table = 'erg_unregulated_facilities'
                unreg_type = 'facilities'
                join_sql = 'a.facility_id = b.facility_id'
                detail_cols = 'a.facility_id'
            elif view_has_column('tank_id'):
                unreg_table = 'erg_unregulated_tanks'
                unreg_type = 'tanks'
                join_sql = 'a.facility_id = b.facility_id and a.tank_id = b.tank_id'
                detail_cols = 'a.facility_id, a.tank_id'
            else:
                unreg_table = 'erg_unregulated_facilities'
                unreg_type = 'facilities'
                join_sql = 'a.facility_id = b.facility_id'
                detail_cols = 'a.facility_id'
        else:
            if not view_has_column('release_id'):
                msg = f'Missing column {self.view_name}.release_id; skipping unregulated parent check.'
                logger.warning(msg)
                self.error_cnt_dict['Rows with unregulated rows not excluded from ' + self.view_name] = 0
                self.error_dict[msg] = 'Skipped'
                return

            unreg_table = 'erg_unregulated_releases'
            unreg_type = 'releases'
            join_sql = 'a.release_id = b.release_id'
            detail_cols = 'a.release_id'

        if not utils.get_table_existence(unreg_table, self.dataset.schema):
            msg = f'Missing table {self.dataset.schema}.{unreg_table}; skipping unregulated {unreg_type} check. Run create-unreg first.'
            logger.warning(msg)
            self.error_cnt_dict['Rows with unregulated ' + unreg_type + ' not excluded from ' + self.view_name] = 0
            self.error_dict[msg] = 'Skipped'
            return

        sql = f"""select count(*)
                  from {self.dataset.schema}.{self.view_name} a join {self.dataset.schema}.{unreg_table} b on {join_sql}"""
        num_rows = self._select_count(sql)
        self.error_cnt_dict['Rows with unregulated ' + unreg_type + ' not excluded from ' + self.view_name] = num_rows
        logger.warning('Rows with unregulated %s in %s: %s', unreg_type, self.view_name, num_rows)
        if num_rows > 0:
            self.error_dict['Unregulated ' + unreg_type + ' in ' + self.view_name] =  num_rows
            if self.include_details:
                unreg_reason_sql = 'b.unregulated_reason'
                if not self._relation_has_column(unreg_table, 'unregulated_reason'):
                    unreg_reason_sql = 'null::varchar(1000) as unregulated_reason'
                detail_sql = f"""select {detail_cols}, {unreg_reason_sql}
                               from {self.dataset.schema}.{self.view_name} a join {self.dataset.schema}.{unreg_table} b on {join_sql}"""
                utils.process_sql(self.conn, self.cur, detail_sql)
                data = self.cur.fetchall()
                self.write_to_ws(data, 'Unreg ' + self.view_name.replace('v_ust_',''))


    def check_missing_parent_view_keys(self):
        parent_specs_ust = {
            'v_ust_tank': ('v_ust_facility', ['facility_id']),
            'v_ust_tank_substance': ('v_ust_tank', ['facility_id', 'tank_id']),
            'v_ust_tank_dispenser': ('v_ust_tank', ['facility_id', 'tank_id']),
            'v_ust_compartment': ('v_ust_tank', ['facility_id', 'tank_id']),
            'v_ust_compartment_substance': ('v_ust_compartment', ['facility_id', 'tank_id', 'compartment_id']),
            'v_ust_compartment_dispenser': ('v_ust_compartment', ['facility_id', 'tank_id', 'compartment_id']),
            'v_ust_piping': ('v_ust_compartment', ['facility_id', 'tank_id', 'compartment_id']),
        }
        parent_specs_release = {
            'v_ust_release_source': ('v_ust_release', ['release_id']),
            'v_ust_release_cause': ('v_ust_release', ['release_id']),
            'v_ust_release_substance': ('v_ust_release', ['release_id']),
            'v_ust_release_corrective_action_strategy': ('v_ust_release', ['release_id']),
        }

        if self.dataset.ust_or_release == 'ust':
            spec = parent_specs_ust.get(self.view_name)
        else:
            spec = parent_specs_release.get(self.view_name)

        if not spec:
            return

        parent_view, key_cols = spec
        error_label = (
            'Rows in ' + self.dataset.schema + '.' + self.view_name
            + ' missing keys in ' + self.dataset.schema + '.' + parent_view
        )
        if not self.include_details:
            logger.info('Skipping missing parent key check for %s.%s in fast QA mode.', self.dataset.schema, self.view_name)
            self.error_cnt_dict[error_label] = 0
            return

        child_cols = set(self._get_view_columns(self.view_name))
        missing_child_key_cols = [col for col in key_cols if col not in child_cols]
        if missing_child_key_cols:
            msg = f'Missing key columns in {self.view_name}; skipping parent key check: {", ".join(missing_child_key_cols)}'
            logger.warning(msg)
            self.error_dict[msg] = 'Skipped'
            self.error_cnt_dict[error_label] = 0
            return

        if parent_view not in self.views_to_review:
            msg = f'Missing parent view {self.dataset.schema}.{parent_view}; skipping parent key check for {self.view_name}.'
            logger.warning(msg)
            self.error_dict[msg] = 'Skipped'
            self.error_cnt_dict[error_label] = 0
            return

        if not utils.get_table_existence(parent_view, self.dataset.schema):
            msg = f'Parent view {self.dataset.schema}.{parent_view} not found; skipping parent key check for {self.view_name}.'
            logger.warning(msg)
            self.error_dict[msg] = 'Skipped'
            self.error_cnt_dict[error_label] = 0
            return

        key_col_sql = ', '.join(self._quote_ident(col) for col in key_cols)
        missing_key_sql = f"""select distinct {key_col_sql}
                              from {self.dataset.schema}.{self.view_name}
                              except
                              select distinct {key_col_sql}
                              from {self.dataset.schema}.{parent_view}"""
        missing_key_join_sql = ' and '.join(
            f'a.{self._quote_ident(col)} = missing_keys.{self._quote_ident(col)}'
            for col in key_cols
        )
        missing_parent_sql = f"""with missing_keys as (
                  {missing_key_sql}
                  )
                  select count(*)
                  from {self.dataset.schema}.{self.view_name} a
                  join missing_keys on {missing_key_join_sql}"""
        num_rows = self._select_count(missing_parent_sql)
        self.error_cnt_dict[error_label] = num_rows
        logger.warning('%s: %s', error_label, num_rows)

        if num_rows > 0:
            self.error_dict['Missing parent keys for ' + self.view_name + ' in ' + parent_view] = num_rows
            if self.include_details:
                order_sql = ', '.join(self._quote_ident(col) for col in key_cols)
                detail_sql = f"""{missing_key_sql}
                         order by {order_sql}"""
                utils.process_sql(self.conn, self.cur, detail_sql)
                data = self.cur.fetchall()
                ws_name = ('Missing parent ' + self.view_name.replace('v_ust_', ''))[:31]
                ws = self.wb.create_sheet(ws_name)
                ws.append(key_cols)
                for row in data:
                    ws.append(list(row))


    def check_compartment_data_flag(self):
        sql = "select organization_compartment_flag from public.ust_control where ust_control_id = %s"
        utils.process_sql(self.conn, self.cur, sql, params=(self.dataset.control_id,))
        org_comp_flag = self.cur.fetchone()[0]
        if not org_comp_flag:
            self.error_dict['Missing organization_compartment_flag in ust_control'] = org_comp_flag 
            logger.warning('Missing organization_compartment_flag in ust_control')
        elif org_comp_flag not in ['Y','N']:
            self.error_dict['Bad value of in organization_compartment_flag in ust_control'] = org_comp_flag
            logger.warning('Bad value of %s in organization_compartment_flag in ust_control', org_comp_flag)


    def write_overview(self):
        # Print an overview of QA/QC results
        ws = self.wb.create_sheet('Overview')
        rowno = 1
        ws.cell(row=rowno, column=1).value = 'View Name'
        ws.cell(row=rowno, column=1).font = Font(bold=True)
        ws.cell(row=rowno, column=2).value = 'Number of Rows'
        ws.cell(row=rowno, column=2).font = Font(bold=True)
        rowno +=1 
        for k, v in self.view_counts.items():
            print('Number of rows in ' + k + ' = ' + str(v))
            ws.cell(row=rowno, column=1).value = k
            ws.cell(row=rowno, column=2).value = v  
            rowno += 1

        rowno += 2
        ws.cell(row=rowno, column=1).value = 'QA Check'
        ws.cell(row=rowno, column=1).font = Font(bold=True)
        ws.cell(row=rowno, column=2).value = 'Number of Rows'
        ws.cell(row=rowno, column=2).font = Font(bold=True)
        rowno +=1     
        for k, v in self.error_cnt_dict.items():
            print(k + ' = ' + str(v))
            ws.cell(row=rowno, column=1).value = k
            ws.cell(row=rowno, column=2).value = v  
            if v > 0:
                ws.cell(row=rowno, column=2).fill = utils.get_fill_gen(yellow_cell_fill)
            rowno += 1
    
        utils.autowidth(ws)        
    
        rowno += 2
        ws.cell(row=rowno, column=1).value = 'Bad or Missing Data'
        ws.cell(row=rowno, column=1).font = Font(bold=True)
        ws.cell(row=rowno, column=2).value = 'Details'
        ws.cell(row=rowno, column=2).font = Font(bold=True)
        rowno +=1     
        bad = False
        for k, v in self.error_dict.items():
            bad = True
            print(k + ' = ' + str(v))
            ws.cell(row=rowno, column=1).value = k
            ws.cell(row=rowno, column=2).value = v  
            rowno += 1
        if not bad: 
            ws.cell(row=rowno, column=1).value = 'No bad or missing data'
            ws.cell(row=rowno, column=1).font = Font(italic=True)


    def exclusions(self):
        if self.error_dict and not self.force_exclusions:
            return 
        exclusions = Exclusions(self.dataset).exclusions
        for tab_name, metadata in exclusions.items():
            logger.info('Working on "%s"', tab_name)
            ws = self.wb.create_sheet(tab_name)
            rowno = 1        
            colno = 1
            for header in metadata['headers']:
                ws.cell(row=rowno, column=colno).value = header
                ws.cell(row=rowno, column=colno).font = Font(bold=True)
                colno += 1
            rowno +=1 
            for row in metadata['data']: 
                colno = 1
                for col in row:                
                    ws.cell(row=rowno, column=colno).value = row[colno-1]
                    colno += 1
                rowno += 1
            utils.autowidth(ws)        
            utils.add_ws_filter(ws)
        logger.info('Added exclusion tabs')


    def summary_counts(self):
        if self.error_dict and not self.force_summary_counts:
            return 
        summ_counts = SummaryCounts(self.dataset).summ_counts 

        for k, rows in summ_counts.items():
            logger.info('Working on "%s"', k)
            ws = self.wb.create_sheet(k)
            rowno = 1
            ws.cell(row=rowno, column=1).value = 'EPA Value'
            ws.cell(row=rowno, column=1).font = Font(bold=True)
            ws.cell(row=rowno, column=2).value = 'Number of Rows'
            ws.cell(row=rowno, column=2).font = Font(bold=True)
            rowno +=1 
            for row in rows:
                print(row[0] + ' = ' + str(row[1]))
                ws.cell(row=rowno, column=1).value = row[0]
                ws.cell(row=rowno, column=2).value = row[1]  
                rowno += 1
            utils.autowidth(ws)        
    
        logger.info('Added summary count tabs')


    def cleanup_wb(self):
        try:
            self.wb.remove(self.wb['Sheet'])
            self.wb.active = self.wb['Overview']
        except KeyError:
            pass
        self.wb.save(self.dataset.export_file_path)



def main(ust_or_release, 
         control_id=0, 
         organization_id=None, 
         force_exclusions=False,
         force_summary_counts=False, 
         include_details=True,
         export_file_name=None, 
         export_file_dir=None, 
         export_file_path=None):
    if not control_id or control_id == 0:
        control_id = utils.get_control_id(ust_or_release, organization_id.upper())

    dataset = Dataset(ust_or_release=ust_or_release,
                      control_id=control_id, 
                      base_file_name='QAQC_' + utils.get_timestamp_str() + '.xlsx',
                      export_file_name=export_file_name,
                      export_file_dir=export_file_dir,
                      export_file_path=export_file_path)

    qc = QualityCheck(
        dataset=dataset,
        force_exclusions=force_exclusions,
        force_summary_counts=force_summary_counts,
        include_details=include_details,
    )
    qc.process()

if __name__ == '__main__':   
    main(ust_or_release=ust_or_release,
         control_id=control_id, 
         organization_id=organization_id,
         force_exclusions=force_exclusions,
         force_summary_counts=force_summary_counts,
         include_details=include_details,
         export_file_name=export_file_name,
         export_file_dir=export_file_dir,
         export_file_path=export_file_path)
