import os
from pathlib import Path
import sys  
ROOT_PATH = Path(__file__).parent.parent.parent
sys.path.append(os.path.join(ROOT_PATH, ''))

import openpyxl as op
from openpyxl.styles import Alignment, Font

from python.util import utils, config
from python.util.dataset import Dataset 
from python.util.logger_factory import logger


ust_or_release = 'ust' 			# Valid values are 'ust' or 'release'
control_id = 0                  # Enter an integer that is the ust_control_id or release_control_id



def build_ws(dataset, ws, admin=False):
	logger.info('Working on Element Mapping tab')
	ws.title = 'Element Mapping'

	if admin:
		headers = ['EPA Table Name','EPA Column Name', 'Organization Table Name', 'Organization Column Name',
				   'Element Mapping ID', 'Programmer Comments', 'Query Logic', 'EPA Comments', 'Organization Comments', 'Inferred Value Comment']		
	else:
		headers = ['Table Name', 'Element Name', 'Organization Table Name', 'Organization Column Name', 
		           'Programmer Comments', 'Query Logic', 'EPA Comments', 'Organization Comments', 'Inferred Value Comment']
	for colno, header in enumerate(headers, start=1):
		ws.cell(row=1, column=colno).value = header
	header_range = ws["A1:J1"]
	for row in header_range:
		for cell in row:
			cell.font = Font(bold=True)

	conn = utils.connect_db()
	cur = conn.cursor()
	logger.info('Connected to database')
	
	if admin:
		cols = f'epa_table_name, epa_column_name, organization_table_name, organization_column_name, {dataset.ust_or_release}_element_mapping_id, programmer_comments, query_logic, epa_comments, organization_comments, inferred_value_comment'
	else:
		cols = 'table_name, element_name, organization_table_name, organization_column_name, programmer_comments, query_logic, epa_comments, organization_comments, inferred_value_comment'
	sql = f"""select {cols}
			from public.v_{dataset.ust_or_release}_element_mapping_for_export
			where {dataset.ust_or_release}_control_id = %s
			order by table_sort_order, column_sort_order"""
	utils.process_sql(conn, cur, sql, params=(dataset.control_id,))
	data = cur.fetchall()
	cur.close()
	conn.close()
	logger.info('Disconnected from database')
	
	for rowno, row in enumerate(data, start=2):
		for colno, cell_value in enumerate(row, start=1):
			ws.cell(row=rowno, column=colno).value = cell_value
	utils.autowidth(ws)
	ws.column_dimensions['E'].width = 50
	ws.column_dimensions['F'].width = 50
	ws.column_dimensions['G'].width = 50
	ws.column_dimensions['H'].width = 50

	logger.info('Finshed Element Mapping tab')
	return ws


def main(ust_or_release, control_id):
	dataset = Dataset(ust_or_release=ust_or_release,
				 	  control_id=control_id, 
				 	  base_file_name='element_mapping_' + utils.get_timestamp_str() + '.xlsx')

	wb = op.Workbook()
	wb.save(dataset.export_file_path)
	ws = wb.create_sheet()
	ws = build_ws(dataset, ws)
	wb.remove(wb['Sheet'])
	wb.save(dataset.export_file_path)



if __name__ == '__main__':   
	main(ust_or_release, control_id)