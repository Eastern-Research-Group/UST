from datetime import date
import ntpath
import os
from pathlib import Path
import sys  
ROOT_PATH = Path(__file__).parent.parent.parent
sys.path.append(os.path.join(ROOT_PATH, ''))

import openpyxl as op
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side

from python.util import utils, config
from python.util.logger_factory import logger


gray_cell_fill = 'C9C9C9' # gray
thin_border = Border(left=Side(style='thin'), 
					 right=Side(style='thin'), 
					 top=Side(style='thin'), 
					 bottom=Side(style='thin'))
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)



class RowCounts:	
	def __init__(self):
		self.export_file_name = 'Element_row_counts_' + utils.get_timestamp_str() + '.xlsx'
		self.export_file_dir = '../../python/exports/other/'
		Path(self.export_file_dir).mkdir(parents=True, exist_ok=True)
		self.export_file_path = self.export_file_dir + self.export_file_name
		self.release_orgs = self.get_orgs('release')
		self.ust_orgs = self.get_orgs('ust')
		self.wb = op.Workbook()
		self.process('release')
		# self.process('ust')
		self.cleanup_wb()
		logger.info('Spreadsheet exported to %s', self.export_file_path)


	def cleanup_wb(self):
		try:
			self.wb.remove(self.wb['Sheet'])
		except:
			pass
		self.wb.save(self.export_file_path)


	def get_orgs(self, ust_or_release):
		conn = utils.connect_db()
		cur = conn.cursor()
		sql = f"select organization_id, {ust_or_release}_control_id from public.v_{ust_or_release}_control where epa_tables_populated = 'Y' order by organization_id"
		cur.execute(sql)
		rows = cur.fetchall()
		orgs = {}
		for row in rows:
			orgs[row[0]] = row[1]
		cur.close()
		conn.close()
		return orgs


	def ws_setup(self, tab_name):
		ws = self.wb.create_sheet(tab_name)

		cell = ws.cell(row=1, column=1)
		cell.value = tab_name
		cell.font = Font(bold=True)
		cell = ws.cell(row=2, column=1)
		cell.value = 'State'
		cell.font = Font(bold=True)
		cell = ws.cell(row=3, column=1)
		cell.value = 'Count of rows'
		cell.font = Font(bold=True)

		cell = ws.cell(row=4, column=1)
		cell.value = 'Element Name'
		cell.font = Font(bold=True)
		cell.fill = utils.get_fill_gen(gray_cell_fill)
		cell.border = thin_border
		cell.alignment = center_align

		ws.freeze_panes = ws['A5']
		return ws


	def process(self, ust_or_release):	
		conn = utils.connect_db()
		cur = conn.cursor()		

		sql = f"select view_name, template_tab_name from public.{ust_or_release}_template_data_tables order by sort_order"
		cur.execute(sql)
		rows = cur.fetchall()
		for row in rows:
			view_name = row[0]
			tab_name = row[1]
			logger.info('Working on %s tab', tab_name)
			ws = self.ws_setup(tab_name)

			sql = f"""select column_name from information_schema.columns 
					where table_schema = 'public' and table_name = '{view_name}'
					and column_name not like '%%control_id' order by ordinal_position"""
			cur.execute(sql)
			rows2 = cur.fetchall()
			rowno = 5
			for row2 in rows2:
				column_name = row2[0]
				logger.info('Working on element %s', column_name)
				cell = ws.cell(row=rowno, column=1)
				cell.value = column_name
				cell.font = Font(bold=True)

				if ust_or_release == 'ust':
					orgs = self.ust_orgs
				else:
					orgs = self.release_orgs

				colno = 2
				for org_id, control_id in orgs.items():
					logger.info('Working on %s: %s', org_id, control_id)

					cell = ws.cell(row=2, column=colno)
					cell.value = org_id
					cell.font = Font(bold=True)

					sql = f"select count(*) from public.{view_name} where {ust_or_release}_control_id = %s"
					cur.execute(sql, (control_id,))
					total_rows = cur.fetchone()[0]
					cell = ws.cell(row=3, column=colno)
					cell.value = total_rows
					cell.font = Font(bold=True)

					cell = ws.cell(row=4, column=colno)
					cell.value = 'Count of non-nulls'
					cell.font = Font(bold=True)
					cell.fill = utils.get_fill_gen(gray_cell_fill)
					cell.border = thin_border
					cell.alignment = center_align

					sql = f"""select count(*) from public.{view_name} where {ust_or_release}_control_id = %s and "{column_name}" is not null"""
					cur.execute(sql, (control_id,))
					num_rows = cur.fetchone()[0]
					cell = ws.cell(row=rowno, column=colno)
					cell.value = num_rows
					
					colno += 1
				rowno += 1

			utils.autowidth(ws)

		cur.close()
		conn.close()




def main():
	row_counts = RowCounts()


if __name__ == '__main__':   
	main()