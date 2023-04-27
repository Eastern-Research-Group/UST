from logger_factory import logger
import utils
import config

from datetime import date
import openpyxl as op
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side


def get_headers(cur, view_name):
	sql = """select column_name from information_schema.columns
	         where table_schema = 'public' and table_name = %s
			 order by ordinal_position"""
	cur.execute(sql, (view_name, ))
	return [x[0] for x in cur.fetchall()]


def delete_col_by_heading(ws, heading):
	i = 1
	for cell in ws[1:1]:  
		if cell.value == heading:
			ws.delete_cols(i)
			return
		i += 1

def make_data_page(state, ust_or_lust, wb):
	ws_name = state + ' ' + ust_or_lust.upper()
	ws = wb.active
	ws.title = ws_name
	
	conn = utils.connect_db()
	cur = conn.cursor()

	view_name = 'v_' + ust_or_lust.lower() 
	headers = get_headers(cur, view_name)

	for colno, header in enumerate(headers, start=1):
		cell = ws.cell(row=1, column=colno)
		cell.value = header
		cell.font = Font(bold=True)

	sql = "select * from public." + view_name + ' where state = %s'
	if ust_or_lust.lower() == 'ust':
		sql = sql + ' order by "FacilityID", "TankID", "CompartmentID"'
	else:
		sql = sql + ' order by "FacilityID", "SiteName", "LUSTID"'
	cur.execute(sql, (state, ))
	data = cur.fetchall()

	for rowno, row in enumerate(data, start=2):
		for colno, cell_value in enumerate(row, start=1):
			ws.cell(row=rowno, column=colno).value = cell_value

	cur.close()
	conn.close()

	ws.delete_cols(1)
	delete_col_by_heading(ws, 'gc_coordinate_source')
	delete_col_by_heading(ws, 'gc_address_type')

	ws.freeze_panes = ws['A2']


def main(state, ust_or_lust, data_only=False):
	file_path = config.local_ust_path + state.upper() + '/'
	file_name = state.upper() + '_' + ust_or_lust.upper() + '_for_geoprocessing-' + str(date.today()) + '.xlsx'
	path = file_path + file_name
	
	wb = op.Workbook()
	make_data_page(state, ust_or_lust, wb)

	wb.save(path)


if __name__ == '__main__':   
	state = 'TN'
	ust_or_lust = 'lust'
	main(state, ust_or_lust)